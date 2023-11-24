#!/usr/bin/env python3

import copy
import glob
import json
import os
import os.path as os_path
import re
import subprocess
import sys
import textwrap
from typing import Dict, Generator, Iterable, List, Literal, Optional, Set, Tuple, Union

from PySide6.QtCore import (
    QElapsedTimer,
    QModelIndex,
    QObject,
    QPersistentModelIndex,
    QPoint,
    Qt,
    QThread,
    QTime,
    QTimer,
    Signal,
)
from PySide6.QtGui import (
    QAction,
    QBrush,
    QColor,
    QCursor,
    QPainter,
    QPalette,
    QStandardItem,
    QStandardItemModel,
    QTextBlockFormat,
    QTextCursor,
)
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFontDialog,
    QGridLayout,
    QGroupBox,
    QHeaderView,
    QLabel,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QSizePolicy,
    QSpacerItem,
    QSplitter,
    QStyledItemDelegate,
    QStyleOptionViewItem,
    QTableView,
    QTabWidget,
    QTextEdit,
    QWidget,
)

from .neosca.lca.lca import LCA
from .neosca.neosca import NeoSCA
from .neosca.structure_counter import StructureCounter
from .ng_about import __name__, __version__
from .ng_io import SCAIO
from .ng_qss import Ng_QSS
from .ng_settings_default import (
    DEFAULT_FONT_FAMILY,
    DEFAULT_FONT_SIZE,
    DEFAULT_INTERFACE_SCALING,
    settings_default,
)


class Ng_Model(QStandardItemModel):
    data_cleared = Signal()
    # itemChanged ^ !data_cleared
    data_updated = Signal()
    data_exported = Signal()

    def __init__(self, *args, main, orientation: Literal["hor", "ver"] = "hor", **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.main = main
        self.orientation = orientation

        self.has_been_exported: bool = False
        self.data_exported.connect(lambda: self.set_has_been_exported(True))
        self.data_updated.connect(lambda: self.set_has_been_exported(False))

    def set_item_str(self, rowno: int, colno: int, value: Union[QStandardItem, str]) -> None:
        item = value if isinstance(value, QStandardItem) else QStandardItem(value)
        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.setItem(rowno, colno, item)

    def set_row_str(self, rowno: int, values: Iterable[Union[QStandardItem, str]]) -> None:
        for colno, value in enumerate(values):
            self.set_item_str(rowno, colno, value)

    def set_item_num(self, rowno: int, colno: int, value: Union[QStandardItem, int, float, str]) -> None:
        if isinstance(value, QStandardItem):
            item = value
        else:
            if not isinstance(value, str):
                value = str(value)
            item = QStandardItem(value)
        item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.setItem(rowno, colno, item)

    def set_row_num(self, rowno: int, values: Iterable[Union[QStandardItem, int, float, str]]) -> None:
        for colno, value in enumerate(values):
            self.set_item_num(rowno, colno, value)

    def set_has_been_exported(self, exported: bool) -> None:
        self.has_been_exported = exported

    def _clear_data(self, leave_an_empty_record=True) -> None:
        if self.orientation == "hor":
            self.setRowCount(0)
            if leave_an_empty_record:
                self.setRowCount(1)
        elif self.orientation == "ver":
            self.setColumnCount(0)
            if leave_an_empty_record:
                self.setColumnCount(1)
        self.data_cleared.emit()

    def clear_data(self, confirm=False, leave_an_empty_record=True) -> None:
        """
        Clear data, reserve headers
        """
        if not confirm or self.has_been_exported:
            return self._clear_data(leave_an_empty_record=leave_an_empty_record)

        messagebox = QMessageBox(self.main)
        messagebox.setWindowTitle("Clear Table")
        messagebox.setText("The table has not been exported yet and all the data will be lost. Continue?")
        messagebox.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        messagebox.accepted.connect(lambda: self._clear_data(leave_an_empty_record=leave_an_empty_record))
        messagebox.exec()

    def is_empty(self):
        for row in range(self.rowCount()):
            for column in range(self.columnCount()):
                item = self.item(row, column)
                if item is not None and item.text() != "":
                    return False
        return True

    def remove_single_empty_row(self) -> None:
        if self.rowCount() == 1 and self.item(0, 0) is None:
            self.setRowCount(0)

    # https://stackoverflow.com/questions/75038194/qt6-how-to-disable-selection-for-empty-cells-in-qtableview
    def flags(self, index) -> Qt.ItemFlag:
        if index.data() is None:
            return Qt.ItemFlag.NoItemFlags
        return super().flags(index)


class Ng_Delegate_SCA(QStyledItemDelegate):
    def __init__(self, parent=None, qss: str = ""):
        super().__init__(parent)
        if (
            triangle_rgb := Ng_QSS.get_value(qss, "QHeaderView::section:horizontal", "background-color")
        ) is not None:
            self.triangle_rgb = triangle_rgb
        else:
            self.triangle_rgb = "#000000"

        self.pos_dialog_mappings: Dict[Tuple[int, int], Ng_Dialog_Text_Edit_SCA_Matched_Subtrees] = {}

    @staticmethod
    def is_index_clickable(index) -> bool:
        data_in_user_role = index.data(Qt.ItemDataRole.UserRole)
        return data_in_user_role is not None and data_in_user_role

    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index: QModelIndex):
        super().paint(painter, option, index)
        if self.is_index_clickable(index):
            painter.save()
            # painter.setBrush(QBrush(Qt.GlobalColor.darkGray))
            painter.setBrush(QBrush(QColor.fromString(self.triangle_rgb)))
            painter.drawPolygon(
                (
                    QPoint(option.rect.x() + option.fontMetrics.descent() * 2, option.rect.y()),
                    QPoint(option.rect.x(), option.rect.y()),
                    QPoint(option.rect.x(), option.rect.y() + option.fontMetrics.descent() * 2),
                )
            )
            painter.restore()

    def createEditor(self, parent, option, index):
        if not self.is_index_clickable(index):
            return None
        pos = (index.row(), index.column())
        if pos in self.pos_dialog_mappings:
            self.pos_dialog_mappings[pos].activateWindow()
        else:
            dialog = Ng_Dialog_Text_Edit_SCA_Matched_Subtrees(parent, index=index)
            self.pos_dialog_mappings[pos] = dialog
            dialog.finished.connect(lambda: self.pos_dialog_mappings.pop(pos))
            dialog.show()


class Ng_TableView(QTableView):
    def __init__(
        self,
        *args,
        main,
        model: Ng_Model,
        has_horizontal_header: bool = True,
        has_vertical_header: bool = True,
        **kwargs,
    ) -> None:
        super().__init__(*args, **kwargs)
        self.main = main
        self.setModel(model)
        self.model_: Ng_Model = model
        self.model_.data_cleared.connect(self.on_data_cleared)
        self.model_.data_updated.connect(self.on_data_updated)
        self.has_horizontal_header = has_horizontal_header
        self.has_vertical_header = has_vertical_header

        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.horizontalHeader().setHighlightSections(False)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.verticalHeader().setHighlightSections(False)
        self.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)

        if self.model_.is_empty():
            self.setEnabled(False)

    def on_data_cleared(self) -> None:
        self.setEnabled(False)

    def on_data_updated(self) -> None:
        # TODO: only need to enable at the first time
        if not self.isEnabled():
            self.setEnabled(True)
        self.scrollToBottom()

    def model(self) -> Ng_Model:
        """Override QTableView().model()"""
        return self.model_

    def set_openpyxl_horizontal_header_alignment(self, cell) -> None:
        from openpyxl.styles import Alignment

        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def set_openpyxl_vertical_header_alignment(self, cell) -> None:
        from openpyxl.styles import Alignment

        if self.has_vertical_header:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def set_openpyxl_cell_alignment(self, cell, item: QStandardItem) -> None:
        # https://doc.qt.io/qtforpython-6/PySide6/QtCore/Qt.html#PySide6.QtCore.PySide6.QtCore.Qt.AlignmentFlag
        # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.alignment.html
        # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L887
        # Qt
        #  Horizontal: Qt.AlignLeft, Qt.AlignRight,  Qt.AlignHCenter, Qt.AlignJustify
        #  Vertical:   Qt.AlignTop,  Qt.AlignBottom, Qt.AlignVCenter, Qt.AlignBaseline
        # OpenPyXL
        #  Horizontal: justify, center, distributed, left, right, fill, general, centerContinuous
        #  Vertical:   justify, center, distributed, top,  bottom

        from openpyxl.styles import Alignment

        alignment_item: Qt.AlignmentFlag = item.textAlignment()

        # Horizontal
        if alignment_item & Qt.AlignmentFlag.AlignLeft:
            alignment_cell_horizontal = "left"
        elif alignment_item & Qt.AlignmentFlag.AlignRight:
            alignment_cell_horizontal = "right"
        elif alignment_item & Qt.AlignmentFlag.AlignHCenter:
            alignment_cell_horizontal = "center"
        elif alignment_item & Qt.AlignmentFlag.AlignJustify:
            alignment_cell_horizontal = "justify"
        else:
            alignment_cell_horizontal = "left"

        # Vertical
        if Qt.AlignmentFlag.AlignTop in alignment_item:
            alignment_cell_vertical = "top"
        elif Qt.AlignmentFlag.AlignBottom in alignment_item:
            alignment_cell_vertical = "bottom"
        elif Qt.AlignmentFlag.AlignVCenter in alignment_item:
            alignment_cell_vertical = "center"
        # > Wordless: Not sure
        elif Qt.AlignmentFlag.AlignBaseline in alignment_item:
            alignment_cell_vertical = "justify"
        else:
            alignment_cell_vertical = "center"

        cell.alignment = Alignment(
            horizontal=alignment_cell_horizontal, vertical=alignment_cell_vertical, wrap_text=True
        )

    def export_table(self) -> None:
        file_path, file_type = QFileDialog.getSaveFileName(
            parent=None,
            caption="Export Table",
            dir=os_path.normpath(os_path.expanduser("~/Desktop")),
            filter="Excel Workbook (*.xlsx);;CSV File (*.csv);;TSV File (*.tsv)",
        )
        if not file_path:
            return

        model: Ng_Model = self.model()
        col_count = model.columnCount()
        row_count = model.rowCount()
        try:
            if ".xlsx" in file_type:
                # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L701C1-L716C54
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                from openpyxl.utils import get_column_letter

                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet_cell = worksheet.cell

                rowno_cell_offset = 2 if self.has_horizontal_header else 1
                colno_cell_offset = 2 if self.has_vertical_header else 1

                # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L628C3-L629C82
                dpi_horizontal = QApplication.primaryScreen().logicalDotsPerInchX()
                dpi_vertical = QApplication.primaryScreen().logicalDotsPerInchY()

                # 1. Horizontal header text and alignment
                if self.has_horizontal_header:
                    for colno_cell, colno_item in enumerate(range(col_count)):
                        cell = worksheet_cell(1, colno_cell_offset + colno_cell)
                        cell.value = model.horizontalHeaderItem(colno_item).text()
                        self.set_openpyxl_horizontal_header_alignment(cell)
                # 2. Vertical header text and alignment
                if self.has_vertical_header:
                    for rowno_cell, rowno_item in enumerate(range(row_count)):
                        cell = worksheet_cell(rowno_cell_offset + rowno_cell, 1)
                        cell.value = model.verticalHeaderItem(rowno_item).text()
                        self.set_openpyxl_vertical_header_alignment(cell)

                # 3. Both header background and font
                # 3.0.1 Get header background
                horizon_bacolor: Optional[str] = Ng_QSS.get_value(
                    self.main.styleSheet(), "QHeaderView::section:horizontal", "background-color"
                )
                vertical_bacolor: Optional[str] = Ng_QSS.get_value(
                    self.main.styleSheet(), "QHeaderView::section:vertical", "background-color"
                )
                # 3.0.2 Get header font, currently only consider color and boldness
                #  https://www.codespeedy.com/change-font-color-of-excel-cells-using-openpyxl-in-python/
                #  https://doc.qt.io/qt-6/stylesheet-reference.html#font-weight
                font_color = Ng_QSS.get_qss_value(self.main.styleSheet(), "QHeaderView::section", "color")
                font_color = font_color.lstrip("#") if font_color is not None else "000"
                font_weight = Ng_QSS.get_qss_value(
                    self.main.styleSheet(), "QHeaderView::section", "font-weight"
                )
                is_bold = (font_weight == "bold") if font_weight is not None else False
                # 3.1 Horizontal header background and font
                if self.has_horizontal_header:
                    # 3.1.1 Horizontal header background
                    #  TODO: Currently all tabpages share the same style sheet and the
                    #   single QSS file is loaded from MainWindow, thus here the
                    #   style sheet is accessed from self. In the future different
                    #   tabs might load their own QSS files, and the style sheet
                    #   should be accessed from the QTabWidget. This is also the
                    #   case for all other "self.styleSheet()" expressions, when
                    #   making this change, remember to edit all of them.
                    if horizon_bacolor is not None:
                        horizon_bacolor = horizon_bacolor.lstrip("#")
                        for colno in range(col_count):
                            cell = worksheet_cell(1, colno_cell_offset + colno)
                            cell.fill = PatternFill(fill_type="solid", fgColor=horizon_bacolor)
                    # 3.1.2 Horizontal header font
                    for colno in range(col_count):
                        cell = worksheet_cell(1, colno_cell_offset + colno)
                        cell.font = Font(color=font_color, bold=is_bold)
                # 3.2 Vertical header background and font
                if self.has_vertical_header:
                    # 3.2.1 Vertial header background
                    if vertical_bacolor is not None:
                        vertical_bacolor = vertical_bacolor.lstrip("#")
                        for rowno in range(row_count):
                            cell = worksheet_cell(rowno_cell_offset + rowno, 1)
                            cell.fill = PatternFill(fill_type="solid", fgColor=vertical_bacolor)
                    # 3.2.2 Vertical header font
                    for rowno in range(row_count):
                        cell = worksheet_cell(rowno_cell_offset + rowno, 1)
                        cell.font = Font(color=font_color, bold=is_bold)

                # 4. Cells
                for rowno in range(row_count):
                    for colno in range(col_count):
                        cell = worksheet_cell(rowno_cell_offset + rowno, colno_cell_offset + colno)
                        item = model.item(rowno, colno)
                        item_value = item.text()
                        try:  # noqa: SIM105
                            item_value = float(item_value)
                        except ValueError:
                            pass
                        cell.value = item_value
                        self.set_openpyxl_cell_alignment(cell, item)
                # 5. Column width
                for colno in range(col_count):
                    # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L729
                    worksheet.column_dimensions[get_column_letter(colno_cell_offset + colno)].width = (
                        self.horizontalHeader().sectionSize(colno) / dpi_horizontal * 13 + 3
                    )

                if self.has_vertical_header:
                    # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L731
                    worksheet.column_dimensions[get_column_letter(1)].width = (
                        self.verticalHeader().width() / dpi_horizontal * 13 + 3
                    )
                # 6. Row height
                worksheet.row_dimensions[1].height = self.horizontalHeader().height() / dpi_vertical * 72
                # 7. Freeze panes
                # https://stackoverflow.com/questions/73837417/freeze-panes-first-two-rows-and-column-with-openpyxl
                # Using "2" in both cases means to always freeze the 1st column
                if self.has_horizontal_header:
                    worksheet.freeze_panes = "B2"
                else:
                    worksheet.freeze_panes = "A2"
                workbook.save(file_path)
            elif ".csv" in file_type or ".tsv" in file_type:
                import csv

                dialect = csv.excel if ".csv" in file_type else csv.excel_tab
                with open(os_path.normpath(file_path), "w", newline="", encoding="utf-8") as fh:
                    csv_writer = csv.writer(fh, dialect=dialect, lineterminator="\n")
                    # Horizontal header
                    hor_header: List[str] = [""]
                    hor_header.extend(model.horizontalHeaderItem(colno).text() for colno in range(col_count))
                    csv_writer.writerow(hor_header)
                    # Vertical header + cells
                    for rowno in range(row_count):
                        row: List[str] = [model.verticalHeaderItem(rowno).text()]
                        row.extend(model.item(rowno, colno).text() for colno in range(col_count))
                        csv_writer.writerow(row)
            QMessageBox.information(
                self, "Success", f"The table has been successfully exported to {file_path}."
            )
        except PermissionError:
            QMessageBox.critical(
                self,
                "Error Exporting Cells",
                f"PermissionError: failed to export the table to {file_path}.",
            )
        else:
            model.data_exported.emit()


# https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_dialogs/wl_dialogs.py#L28
class Ng_Dialog(QDialog):
    def __init__(
        self, *args, title: str = "", width: int = 0, height: int = 0, resizable=False, **kwargs
    ) -> None:
        """
        ┌———————————┐
        │           │
        │  content  │
        │           │
        │———————————│
        │  buttons  │
        └———————————┘
        """
        super().__init__(*args, **kwargs)
        # > Dialog size
        if resizable:
            if not width:
                width = self.size().width()

            if not height:
                height = self.size().height()

            self.resize(width, height)
        else:
            if width:
                self.setFixedWidth(width)

            if height:
                self.setFixedHeight(height)
            # Gives the window a thin dialog border on Windows. This style is
            # traditionally used for fixed-size dialogs.
            self.setWindowFlag(Qt.WindowType.MSWindowsFixedSizeDialogHint)
        self.setWindowTitle(title)

        self.content_layout = QGridLayout()
        self.button_layout = QGridLayout()

        self.grid_layout = QGridLayout()
        self.grid_layout.addLayout(self.content_layout, 0, 0)
        self.grid_layout.addLayout(self.button_layout, 1, 0)
        self.setLayout(self.grid_layout)

    def rowCount(self) -> int:
        return self.content_layout.rowCount()

    def columnCount(self) -> int:
        return self.content_layout.columnCount()

    def addWidget(self, *args, **kwargs) -> None:
        self.content_layout.addWidget(*args, **kwargs)

    def addButton(self, *args, **kwargs) -> None:
        self.button_layout.addWidget(*args, **kwargs)

    def setColumnStretch(self, column: int, strech: int) -> None:
        self.content_layout.setColumnStretch(column, strech)

    def setRowStretch(self, row: int, strech: int) -> None:
        self.content_layout.setRowStretch(row, strech)


class Ng_Dialog_Processing_With_Elapsed_Time(Ng_Dialog):
    started = Signal()
    # Use this to get the place holder, e.g. 0:00:00
    time_format_re = re.compile(r"[^:]")

    def __init__(
        self,
        *args,
        title: str = "Please Wait",
        width: int = 500,
        height: int = 0,
        time_format: str = "h:mm:ss",
        interval: int = 1000,
        **kwargs,
    ) -> None:
        super().__init__(*args, title=title, width=width, height=height, resizable=False, **kwargs)
        self.time_format = time_format
        self.interval = interval
        self.elapsedtimer = QElapsedTimer()
        self.timer = QTimer()

        # TODO: this label should be exposed
        self.label_status = QLabel("Processing...")
        self.text_time_elapsed_zero = f"Elapsed time: {self.time_format_re.sub('0', time_format)}"
        self.label_time_elapsed = QLabel(self.text_time_elapsed_zero)
        self.label_please_wait = QLabel(
            "Please be patient. The wait time can range from a few seconds to several minutes."
        )
        self.label_please_wait.setWordWrap(True)

        self.addWidget(self.label_status, 0, 0)
        self.addWidget(self.label_time_elapsed, 0, 1, alignment=Qt.AlignmentFlag.AlignRight)
        self.addWidget(self.label_please_wait, 1, 0, 1, 2)

        self.setWindowFlags(self.windowFlags() | Qt.WindowType.FramelessWindowHint)

        # Bind
        self.timer.timeout.connect(self.update_time_elapsed)
        self.started.connect(self.elapsedtimer.start)
        # If the timer is already running, it will be stopped and restarted.
        self.started.connect(lambda: self.timer.start(self.interval))
        # Either 'accepted' or 'rejected', although 'rejected' is disabled by
        # overriding the 'reject' method (see below)
        self.finished.connect(self.reset_time_elapsed)
        self.finished.connect(self.timer.stop)

    def reset_time_elapsed(self) -> None:
        self.label_time_elapsed.setText(self.text_time_elapsed_zero)

    def update_time_elapsed(self) -> None:
        time_elapsed: int = self.elapsedtimer.elapsed()
        qtime: QTime = QTime.fromMSecsSinceStartOfDay(time_elapsed)
        self.label_time_elapsed.setText(f"Elapsed time: {qtime.toString(self.time_format)}")

    # Override
    def reject(self) -> None:
        pass

    # Override
    def show(self) -> None:
        self.started.emit()
        return super().show()

    # Override
    def open(self) -> None:
        self.started.emit()
        return super().open()

    # Override
    def exec(self) -> int:
        self.started.emit()
        return super().exec()


class Ng_Dialog_Text_Edit(Ng_Dialog):
    def __init__(self, *args, title: str = "", text: str = "", **kwargs) -> None:
        super().__init__(*args, title=title, resizable=True, **kwargs)
        self.textedit = QTextEdit(text)
        self.textedit.setReadOnly(True)
        # https://stackoverflow.com/questions/74852753/indent-while-line-wrap-on-qtextedit-with-pyside6-pyqt6
        indentation: int = self.fontMetrics().horizontalAdvance(" "*4)
        self.fmt_textedit = QTextBlockFormat()
        self.fmt_textedit.setLeftMargin(indentation)
        self.fmt_textedit.setTextIndent(-indentation)

        self.button_copy = QPushButton("Copy")
        self.button_copy.clicked.connect(self.textedit.selectAll)
        self.button_copy.clicked.connect(self.textedit.copy)

        self.button_close = QPushButton("Close")
        self.button_close.clicked.connect(self.reject)

        self.addButton(self.button_copy, 0, 0, alignment=Qt.AlignmentFlag.AlignLeft)
        self.addButton(self.button_close, 0, 1, alignment=Qt.AlignmentFlag.AlignRight)

    def setText(self, text: str) -> None:
        self.textedit.setText(text)
        cursor = QTextCursor(self.textedit.document())
        cursor.select(QTextCursor.SelectionType.Document)
        cursor.mergeBlockFormat(self.fmt_textedit)

    # Override
    def show(self) -> None:
        # Add self.textedit lastly to allow users add custom widgets above
        self.addWidget(self.textedit, self.rowCount(), 0, 1, self.columnCount())
        return super().show()

    # Override
    def open(self) -> None:
        # Add self.textedit lastly to allow users add custom widgets above
        self.addWidget(self.textedit, self.rowCount(), 0, 1, self.columnCount())
        return super().open()

    # Override
    def exec(self) -> int:
        # Add self.textedit lastly to allow users add custom widgets above
        self.addWidget(self.textedit, self.rowCount(), 0, 1, self.columnCount())
        return super().exec()


class Ng_Dialog_Text_Edit_SCA_Matched_Subtrees(Ng_Dialog_Text_Edit):
    def __init__(self, *args, index: Union[QModelIndex, QPersistentModelIndex], **kwargs):
        super().__init__(*args, title="Matches", width=500, height=300, **kwargs)
        self.file_name = index.model().verticalHeaderItem(index.row()).text()
        self.sname = index.model().horizontalHeaderItem(index.column()).text()
        self.matched_subtrees: List[str] = index.data(Qt.ItemDataRole.UserRole)
        self.setText("\n".join(self.matched_subtrees))

        self.label_summary = QLabel(
            f'{len(self.matched_subtrees)} occurrences of "{self.sname}" in "{self.file_name}"'
        )
        self.label_summary.setWordWrap(True)
        self.addWidget(self.label_summary)


class Ng_Dialog_Text_Edit_Citing(Ng_Dialog_Text_Edit):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, title="Citing", **kwargs)
        # citing.json is at the same dir of __file__
        # TODO: need to have a unified way to get project path.
        with open(os_path.join(os_path.dirname(__file__), "citing.json"), encoding="utf-8") as f:
            self.style_citation_mapping = json.load(f)

        self.label_citing = QLabel(f"If you use {__name__} in your research, please kindly cite as follows.")
        self.label_citing.setWordWrap(True)
        self.setText(next(iter(self.style_citation_mapping.values())))
        self.label_choose_citation_style = QLabel("Choose citation style: ")
        self.combobox_choose_citation_style = QComboBox()
        self.combobox_choose_citation_style.addItems(tuple(self.style_citation_mapping.keys()))
        self.combobox_choose_citation_style.currentTextChanged.connect(
            lambda key: self.setText(self.style_citation_mapping[key])
        )

        self.addWidget(self.label_citing, 0, 0, 1, 2)
        self.addWidget(self.label_choose_citation_style, 1, 0)
        self.addWidget(self.combobox_choose_citation_style, 1, 1)
        self.setColumnStretch(1, 1)


class Ng_Dialog_Table(Ng_Dialog):
    def __init__(
        self,
        *args,
        text: str,
        tableview: Ng_TableView,
        **kwargs,
    ) -> None:
        super().__init__(*args, **kwargs)
        self.tableview: Ng_TableView = tableview

        self.content_layout.addWidget(QLabel(text), 0, 0)
        self.content_layout.addWidget(tableview, 1, 0)

        self.button_ok = QPushButton("OK")
        self.button_ok.clicked.connect(self.accept)
        self.button_export_table = QPushButton("Export table...")
        self.button_export_table.clicked.connect(self.tableview.export_table)
        self.button_layout.addWidget(self.button_export_table, 0, 0, alignment=Qt.AlignmentFlag.AlignLeft)
        self.button_layout.addWidget(self.button_ok, 0, 1, alignment=Qt.AlignmentFlag.AlignRight)


class Ng_Worker(QObject):
    worker_done = Signal()

    def __init__(self, *args, main, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.main = main

    def run(self) -> None:
        raise NotImplementedError()


class Ng_Worker_SCA_Generate_Table(Ng_Worker):
    counter_ready = Signal(StructureCounter, str, int)

    def __init__(self, *args, main, **kwargs) -> None:
        super().__init__(*args, main=main, **kwargs)

    def run(self) -> None:
        input_file_names: Generator[str, None, None] = self.main.yield_added_file_names()
        input_file_paths: Generator[str, None, None] = self.main.yield_added_file_paths()

        sca_kwargs = {
            "is_auto_save": False,
            "odir_matched": "",
            "selected_measures": None,
            "is_reserve_parsed": self.main.checkbox_cache_parsed_trees.isChecked(),
            "is_skip_querying": False,
            "is_skip_parsing": False,
            "config": None,
        }

        attrname = "sca_analyzer"
        try:
            sca_analyzer = getattr(self.main, attrname)
        except AttributeError:
            sca_analyzer = NeoSCA(**sca_kwargs)
            setattr(self.main, attrname, sca_analyzer)
        else:
            sca_analyzer.update_options(sca_kwargs)

        err_file_paths: List[str] = []
        for rowno, (file_name, file_path) in enumerate(zip(input_file_names, input_file_paths)):
            try:
                counter: Optional[StructureCounter] = sca_analyzer.parse_and_query_ifile(file_path)
                # TODO should concern --no-parse, --no-query, ... after adding all available options
            except:
                err_file_paths.append(file_path)
                rowno -= 1
                continue
            if counter is None:
                err_file_paths.append(file_path)
                rowno -= 1
                continue
            self.counter_ready.emit(counter, file_name, rowno)

        if err_file_paths:  # TODO: should show a table
            QMessageBox.information(
                None,
                "Error Processing Files",
                "These files are skipped:\n- {}".format("\n- ".join(err_file_paths)),
            )
        self.worker_done.emit()


class Ng_Worker_LCA_Generate_Table(Ng_Worker):
    def __init__(self, *args, main, **kwargs) -> None:
        super().__init__(*args, main=main, **kwargs)

    def run(self) -> None:
        input_file_names: Generator[str, None, None] = self.main.yield_added_file_names()
        input_file_paths: Generator[str, None, None] = self.main.yield_added_file_paths()

        lca_kwargs = {
            "wordlist": "bnc" if self.main.radiobutton_wordlist_BNC.isChecked() else "anc",
            "tagset": "ud" if self.main.radiobutton_tagset_ud.isChecked() else "ptb",
            "is_stdout": False,
        }
        attrname = "lca_analyzer"
        try:
            lca_analyzer = getattr(self.main, attrname)
        except AttributeError:
            lca_analyzer = LCA(**lca_kwargs)
            setattr(self.main, attrname, lca_analyzer)
        else:
            lca_analyzer.update_options(lca_kwargs)

        err_file_paths: List[str] = []
        model: Ng_Model = self.main.model_lca
        has_trailing_rows: bool = True
        for rowno, (file_name, file_path) in enumerate(zip(input_file_names, input_file_paths)):
            try:
                values = lca_analyzer._analyze(file_path=file_path)
            except:
                err_file_paths.append(file_path)
                rowno -= 1
                continue
            if values is None:  # TODO: should pop up warning window
                err_file_paths.append(file_path)
                rowno -= 1
                continue
            if has_trailing_rows:
                has_trailing_rows = model.removeRows(rowno, model.rowCount() - rowno)
            # Drop file_path
            del values[0]
            model.set_row_num(rowno, values)
            model.setVerticalHeaderItem(rowno, QStandardItem(file_name))
            model.data_updated.emit()

        if err_file_paths:  # TODO: should show a table
            QMessageBox.information(
                None,
                "Error Processing Files",
                "These files are skipped:\n- {}".format("\n- ".join(err_file_paths)),
            )

        self.worker_done.emit()


class Ng_Thread(QThread):
    def __init__(self, worker: Ng_Worker):
        super().__init__()
        self.worker = worker
        # https://mayaposch.wordpress.com/2011/11/01/how-to-really-truly-use-qthreads-the-full-explanation/
        self.worker.moveToThread(self)

    def run(self):
        self.start()
        self.worker.run()

    # def cancel(self) -> None:
    #     self.terminate()
    #     self.wait()


class Ng_Main(QMainWindow):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setup_env()

        self.setWindowTitle(f"{__name__} {__version__}")
        file_path_settings = os_path.join(self.here, "ng_settings.pickle")
        self.settings_custom = SCAIO.load_pickle_file(file_path_settings, None)
        if self.settings_custom is None:
            self.settings_custom = copy.deepcopy(settings_default)
        qss = textwrap.dedent(
            f"""\
            * {{
            font-family: "{self.settings_custom['general']['ui_settings']['font_family']}";
            font-size: {self.settings_custom['general']['ui_settings']['font_size']}pt;
            }}\n"""
        )
        file_path_style_qss = os_path.join(self.here, "ng_style.qss")
        qss += Ng_QSS.read_qss_file(file_path_style_qss, "")
        self.setStyleSheet(qss)
        self.setup_menu()
        self.setup_worker()
        self.setup_main_window()

    def setup_menu(self):
        # File
        self.menu_file = QMenu("File", self.menuBar())
        action_open_file = QAction("Open File...", self.menu_file)
        action_open_file.setShortcut("CTRL+O")
        action_open_file.triggered.connect(self.menubar_file_open_file)
        action_open_folder = QAction("Open Folder...", self.menu_file)
        action_open_folder.setShortcut("CTRL+F")
        action_open_folder.triggered.connect(self.menubar_file_open_folder)
        action_restart = QAction("Restart", self.menu_file)  # TODO remove this before releasing
        action_restart.triggered.connect(self.menubar_file_restart)  # TODO remove this before releasing
        action_restart.setShortcut("CTRL+R")  # TODO remove this before releasing
        action_quit = QAction("Quit", self.menu_file)
        action_quit.setShortcut("CTRL+Q")
        action_quit.triggered.connect(self.close)
        self.menu_file.addAction(action_open_file)
        self.menu_file.addAction(action_open_folder)
        self.menu_file.addSeparator()
        self.menu_file.addAction(action_restart)
        self.menu_file.addAction(action_quit)
        # Preferences
        self.menu_preferences = QMenu("Preferences", self.menuBar())
        action_font = QAction("Font", self.menu_preferences)
        action_font.triggered.connect(self.menubar_preferences_font)
        self.menu_preferences.addAction(action_font)
        # Help
        self.menu_help = QMenu("Help", self.menuBar())
        action_citing = QAction("Citing", self.menu_help)
        action_citing.triggered.connect(self.menubar_help_citing)
        self.menu_help.addAction(action_citing)

        self.menuBar().addMenu(self.menu_file)
        self.menuBar().addMenu(self.menu_preferences)
        self.menuBar().addMenu(self.menu_help)

    def menubar_preferences_font(self) -> None:
        ok, font = QFontDialog.getFont()
        if not ok:
            return
        breakpoint()
        print(ok, font)

    def menubar_help_citing(self) -> None:
        dialog_citing = Ng_Dialog_Text_Edit_Citing(self)
        dialog_citing.exec()

    def setup_tab_sca(self):
        self.button_generate_table_sca = QPushButton("Generate table")
        self.button_generate_table_sca.setShortcut("CTRL+G")
        self.button_export_table_sca = QPushButton("Export all cells...")
        self.button_export_table_sca.setEnabled(False)
        # self.button_export_selected_cells = QPushButton("Export selected cells...")
        # self.button_export_selected_cells.setEnabled(False)
        self.button_clear_table_sca = QPushButton("Clear table")
        self.button_clear_table_sca.setEnabled(False)

        # TODO comment this out before releasing
        self.button_custom_func = QPushButton("Custom func")
        # TODO comment this out before releasing
        self.button_custom_func.clicked.connect(self.custom_func)

        self.checkbox_cache_parsed_trees = QCheckBox("Cache parse trees")
        self.checkbox_cache_parsed_trees.setChecked(True)

        self.model_sca = Ng_Model(main=self)
        self.model_sca.setColumnCount(len(StructureCounter.DEFAULT_MEASURES))
        self.model_sca.setHorizontalHeaderLabels(StructureCounter.DEFAULT_MEASURES)
        self.model_sca.clear_data()
        self.tableview_sca = Ng_TableView(main=self, model=self.model_sca)
        self.tableview_sca.setItemDelegate(Ng_Delegate_SCA(None, self.styleSheet()))

        # Bind
        self.button_generate_table_sca.clicked.connect(self.ng_thread_sca_generate_table.start)
        self.button_export_table_sca.clicked.connect(self.tableview_sca.export_table)
        self.button_export_matches_sca.clicked.connect(self.tableview_sca.export_matches)
        self.button_clear_table_sca.clicked.connect(lambda: self.model_sca.clear_data(confirm=True))
        self.model_sca.data_cleared.connect(
            lambda: self.button_generate_table_sca.setEnabled(True) if not self.model_file.is_empty() else None
        )
        self.model_sca.data_cleared.connect(lambda: self.button_export_table_sca.setEnabled(False))
        self.model_sca.data_cleared.connect(lambda: self.button_clear_table_sca.setEnabled(False))
        self.model_sca.data_updated.connect(lambda: self.button_export_table_sca.setEnabled(True))
        self.model_sca.data_updated.connect(lambda: self.button_clear_table_sca.setEnabled(True))
        self.model_sca.data_updated.connect(lambda: self.button_generate_table_sca.setEnabled(False))

        widget_settings_sca = QWidget()
        layout_settings_sca = QGridLayout()
        widget_settings_sca.setLayout(layout_settings_sca)
        layout_settings_sca.addWidget(self.checkbox_cache_parsed_trees, 0, 0)
        layout_settings_sca.addItem(QSpacerItem(0, 0, vData=QSizePolicy.Policy.Expanding))
        layout_settings_sca.setContentsMargins(1, 0, 1, 0)

        self.scrollarea_settings_sca = QScrollArea()
        self.scrollarea_settings_sca.setWidgetResizable(True)
        self.scrollarea_settings_sca.setBackgroundRole(QPalette.ColorRole.Light)
        self.scrollarea_settings_sca.setMinimumWidth(200)
        self.scrollarea_settings_sca.setWidget(widget_settings_sca)

        self.widget_previewarea_sca = QWidget()
        self.layout_previewarea_sca = QGridLayout()
        self.widget_previewarea_sca.setLayout(self.layout_previewarea_sca)
        for btn_no, btn in enumerate(
            (
                self.button_generate_table_sca,
                self.button_export_table_sca,
                self.button_clear_table_sca,
                self.button_custom_func,
            ),
            start=1,
        ):
            self.layout_previewarea_sca.addWidget(btn, 1, btn_no - 1)
        self.layout_previewarea_sca.addWidget(self.tableview_sca, 0, 0, 1, btn_no)
        self.layout_previewarea_sca.addWidget(self.scrollarea_settings_sca, 0, btn_no, 2, 1)
        self.layout_previewarea_sca.setContentsMargins(0, 0, 0, 0)

        self.splitter_workarea_sca = QSplitter(Qt.Orientation.Horizontal)
        self.splitter_workarea_sca.setChildrenCollapsible(False)
        self.splitter_workarea_sca.addWidget(self.widget_previewarea_sca)
        self.splitter_workarea_sca.addWidget(self.scrollarea_settings_sca)
        self.splitter_workarea_sca.setStretchFactor(0, 5)
        self.splitter_workarea_sca.setStretchFactor(1, 1)
        self.splitter_workarea_sca.setContentsMargins(6, 4, 6, 4)

    def custom_func(self):
        breakpoint()

    def setup_tab_lca(self):
        self.button_generate_table_lca = QPushButton("Generate table")
        self.button_generate_table_lca.setShortcut("CTRL+G")
        self.button_export_table_lca = QPushButton("Export all cells...")
        self.button_export_table_lca.setEnabled(False)
        # self.button_export_selected_cells = QPushButton("Export selected cells...")
        # self.button_export_selected_cells.setEnabled(False)
        self.button_clear_table_lca = QPushButton("Clear table")
        self.button_clear_table_lca.setEnabled(False)

        self.model_lca = Ng_Model(main=self)
        self.model_lca.setColumnCount(len(LCA.FIELDNAMES) - 1)
        self.model_lca.setHorizontalHeaderLabels(LCA.FIELDNAMES[1:])
        self.model_lca.clear_data()
        self.tableview_lca = Ng_TableView(main=self, model=self.model_lca)
        # TODO: tableview_sca use custom delegate to only enable
        # clickable items, in which case a dialog will pop up to show matches.
        # Here when tableview_lca also use custom delegate, remember to
        # remove this line.
        self.tableview_lca.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

        # Bind
        self.button_generate_table_lca.clicked.connect(self.ng_thread_lca_generate_table.start)
        self.button_export_table_lca.clicked.connect(self.tableview_lca.export_table)
        self.button_clear_table_lca.clicked.connect(lambda: self.model_lca.clear_data(confirm=True))
        self.model_lca.data_cleared.connect(
            lambda: self.button_generate_table_lca.setEnabled(True) if not self.model_file.is_empty() else None
        )
        self.model_lca.data_cleared.connect(lambda: self.button_export_table_lca.setEnabled(False))
        self.model_lca.data_cleared.connect(lambda: self.button_clear_table_lca.setEnabled(False))
        self.model_lca.data_updated.connect(lambda: self.button_export_table_lca.setEnabled(True))
        self.model_lca.data_updated.connect(lambda: self.button_clear_table_lca.setEnabled(True))
        self.model_lca.data_updated.connect(lambda: self.button_generate_table_lca.setEnabled(False))

        self.radiobutton_wordlist_BNC = QRadioButton("British National Corpus (BNC) wordlist")
        self.radiobutton_wordlist_BNC.setChecked(True)
        self.radiobutton_wordlist_ANC = QRadioButton("American National Corpus (ANC) wordlist")
        groupbox_wordlist = QGroupBox("Wordlist")
        layout_wordlist = QGridLayout()
        groupbox_wordlist.setLayout(layout_wordlist)
        layout_wordlist.addWidget(self.radiobutton_wordlist_BNC, 0, 0)
        layout_wordlist.addWidget(self.radiobutton_wordlist_ANC, 1, 0)
        self.radiobutton_tagset_ud = QRadioButton("Universal POS Tagset")
        self.radiobutton_tagset_ud.setChecked(True)
        self.radiobutton_tagset_ptb = QRadioButton("Penn Treebank POS Tagset")
        groupbox_tagset = QGroupBox("Tagset")
        layout_tagset = QGridLayout()
        groupbox_tagset.setLayout(layout_tagset)
        layout_tagset.addWidget(self.radiobutton_tagset_ud, 0, 0)
        layout_tagset.addWidget(self.radiobutton_tagset_ptb, 1, 0)

        widget_settings_lca = QWidget()
        layout_settings_lca = QGridLayout()
        widget_settings_lca.setLayout(layout_settings_lca)
        layout_settings_lca.addWidget(groupbox_wordlist, 0, 0)
        layout_settings_lca.addWidget(groupbox_tagset, 1, 0)
        layout_settings_lca.addItem(QSpacerItem(0, 0, vData=QSizePolicy.Policy.Expanding))
        layout_settings_lca.setContentsMargins(1, 0, 1, 0)

        self.scrollarea_settings_lca = QScrollArea()
        self.scrollarea_settings_lca.setWidgetResizable(True)
        self.scrollarea_settings_lca.setBackgroundRole(QPalette.ColorRole.Light)
        self.scrollarea_settings_lca.setMinimumWidth(200)
        self.scrollarea_settings_lca.setWidget(widget_settings_lca)
        
        self.widget_previewarea_lca = QWidget()
        self.layout_previewarea_lca = QGridLayout()
        self.widget_previewarea_lca.setLayout(self.layout_previewarea_lca)
        for btn_no, btn in enumerate(
            (
                self.button_generate_table_lca,
                self.button_export_table_lca,
                self.button_clear_table_lca,
            ),
            start=1,
        ):
            self.layout_previewarea_lca.addWidget(btn, 1, btn_no - 1)
        self.layout_previewarea_lca.addWidget(self.tableview_lca, 0, 0, 1, btn_no)
        self.layout_previewarea_lca.setContentsMargins(0, 0, 0, 0)

        self.splitter_workarea_lca = QSplitter(Qt.Orientation.Horizontal)
        self.splitter_workarea_lca.setChildrenCollapsible(False)
        self.splitter_workarea_lca.addWidget(self.widget_previewarea_lca)
        self.splitter_workarea_lca.addWidget(self.scrollarea_settings_lca)
        self.splitter_workarea_lca.setStretchFactor(0, 5)
        self.splitter_workarea_lca.setStretchFactor(1, 1)
        self.splitter_workarea_lca.setContentsMargins(6, 4, 6, 4)

    def enable_button_generate_table(self, enabled: bool) -> None:
        self.button_generate_table_sca.setEnabled(enabled)
        self.button_generate_table_lca.setEnabled(enabled)

    def setup_tableview_file(self) -> None:
        self.model_file = Ng_Model(main=self)
        self.model_file.setHorizontalHeaderLabels(("Name", "Path"))
        self.model_file.data_cleared.connect(lambda: self.enable_button_generate_table(False))
        self.model_file.data_updated.connect(lambda: self.enable_button_generate_table(True))
        self.model_file.clear_data()
        self.tableview_file = Ng_TableView(main=self, model=self.model_file, has_vertical_header=False)
        self.tableview_file.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tableview_file.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.tableview_file.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tableview_file.setCornerButtonEnabled(True)
        # https://doc.qt.io/qtforpython-6/PySide6/QtWidgets/QWidget.html#PySide6.QtWidgets.PySide6.QtWidgets.QWidget.customContextMenuRequested
        self.menu_tableview_file = QMenu(self)
        self.action_tableview_file_remove = QAction("Remove", self.menu_tableview_file)
        self.action_tableview_file_remove.triggered.connect(self.remove_file_paths)
        self.menu_tableview_file.addAction(self.action_tableview_file_remove)
        self.tableview_file.customContextMenuRequested.connect(self.show_menu_for_tableview_file)

    def setup_main_window(self):
        self.setup_tab_sca()
        self.setup_tab_lca()
        self.setup_tableview_file()

        self.tabwidget = QTabWidget()
        self.tabwidget.addTab(self.splitter_workarea_sca, "Syntactic Complexity Analyzer")
        self.tabwidget.addTab(self.splitter_workarea_lca, "Lexical Complexity Analyzer")
        self.splitter_central_widget = QSplitter(Qt.Orientation.Vertical)
        self.splitter_central_widget.setChildrenCollapsible(False)
        self.splitter_central_widget.addWidget(self.tabwidget)
        self.splitter_central_widget.addWidget(self.tableview_file)
        self.splitter_central_widget.setStretchFactor(0, 2)
        self.splitter_central_widget.setStretchFactor(1, 1)
        self.setCentralWidget(self.splitter_central_widget)

    def sca_add_data(self, counter: StructureCounter, file_name: str, rowno: int) -> None:
        # Remove trailing rows
        self.model_sca.removeRows(rowno, self.model_sca.rowCount() - rowno)
        # Drop file_path
        for colno in range(self.model_sca.columnCount()):
            sname = self.model_sca.horizontalHeaderItem(colno).text()

            value = counter.get_value(sname)
            value_str: str = str(value) if value is not None else ""
            item = QStandardItem(value_str)
            self.model_sca.set_item_num(rowno, colno, item)

            if matches := counter.get_matches(sname):
                item.setData(matches, Qt.ItemDataRole.UserRole)

        self.model_sca.setVerticalHeaderItem(rowno, QStandardItem(file_name))
        self.model_sca.data_updated.emit()

    def setup_worker(self) -> None:
        self.dialog_processing = Ng_Dialog_Processing_With_Elapsed_Time(self)

        self.ng_worker_sca_generate_table = Ng_Worker_SCA_Generate_Table(main=self)
        self.ng_worker_sca_generate_table.counter_ready.connect(self.sca_add_data)
        self.ng_thread_sca_generate_table = Ng_Thread(self.ng_worker_sca_generate_table)
        self.ng_thread_sca_generate_table.started.connect(self.dialog_processing.exec)
        self.ng_thread_sca_generate_table.finished.connect(self.dialog_processing.accept)

        self.ng_worker_lca_generate_table = Ng_Worker_LCA_Generate_Table(main=self)
        self.ng_thread_lca_generate_table = Ng_Thread(self.ng_worker_lca_generate_table)
        self.ng_thread_lca_generate_table.started.connect(self.dialog_processing.exec)
        self.ng_thread_lca_generate_table.finished.connect(self.dialog_processing.accept)

    def setup_env(self) -> None:
        self.here = os_path.dirname(os_path.abspath(__file__))
        ng_home = os_path.dirname(self.here)
        libs_dir = os_path.join(ng_home, "libs")
        # TODO: remove these
        self.java_home = os_path.join(libs_dir, "jdk8u372")
        self.stanford_parser_home = os_path.join(libs_dir, "stanford-parser-full-2020-11-17")
        self.stanford_tregex_home = os_path.join(libs_dir, "stanford-tregex-2020-11-17")
        os.environ["JAVA_HOME"] = self.java_home
        os.environ["STANFORD_PARSER_HOME"] = self.stanford_parser_home
        os.environ["STANFORD_TREGEX_HOME"] = self.stanford_tregex_home
        self.env = os.environ.copy()

    def show_menu_for_tableview_file(self) -> None:
        if not self.tableview_file.selectionModel().selectedRows():
            self.action_tableview_file_remove.setEnabled(False)
        else:
            self.action_tableview_file_remove.setEnabled(True)
        self.menu_tableview_file.exec(QCursor.pos())

    def remove_file_paths(self) -> None:
        # https://stackoverflow.com/questions/5927499/how-to-get-selected-rows-in-qtableview
        indexes: List[QModelIndex] = self.tableview_file.selectionModel().selectedRows()
        # Remove rows from bottom to top, or otherwise lower row indexes will
        # change as upper rows are removed
        rownos = sorted((index.row() for index in indexes), reverse=True)
        for rowno in rownos:
            self.model_file.takeRow(rowno)
        if self.model_file.rowCount() == 0:
            self.model_file.clear_data()

    def remove_model_rows(self, model: Ng_Model, *rownos: int) -> None:
        if not rownos:
            # https://doc.qt.io/qtforpython-6/PySide6/QtGui/QStandardItemModel.html#PySide6.QtGui.PySide6.QtGui.QStandardItemModel.setRowCount
            model.setRowCount(0)
        else:
            for rowno in rownos:
                model.takeRow(rowno)

    # Type hint for generator: https://docs.python.org/3.12/library/typing.html#typing.Generator
    def yield_model_column(self, model: Ng_Model, colno: int) -> Generator[str, None, None]:
        items = (model.item(rowno, colno) for rowno in range(model.rowCount()))
        return (item.text() for item in items if item is not None)

    def yield_added_file_names(self) -> Generator[str, None, None]:
        colno_path = 0
        return self.yield_model_column(self.model_file, colno_path)

    def yield_added_file_paths(self) -> Generator[str, None, None]:
        colno_path = 1
        return self.yield_model_column(self.model_file, colno_path)

    def add_file_paths(self, file_paths_to_add: List[str]) -> None:
        unique_file_paths_to_add: Set[str] = set(file_paths_to_add)
        already_added_file_paths: Set[str] = set(self.yield_added_file_paths())
        file_paths_dup: Set[str] = unique_file_paths_to_add & already_added_file_paths
        file_paths_unsupported: Set[str] = set(
            filter(lambda p: SCAIO.suffix(p) not in SCAIO.SUPPORTED_EXTENSIONS, file_paths_to_add)
        )
        file_paths_empty: Set[str] = set(filter(lambda p: not os_path.getsize(p), unique_file_paths_to_add))
        file_paths_ok: Set[str] = (
            unique_file_paths_to_add
            - already_added_file_paths
            - file_paths_dup
            - file_paths_unsupported
            - file_paths_empty
        )
        if file_paths_ok:
            self.model_file.remove_single_empty_row()
            colno_name = 0
            already_added_file_names = list(
                self.yield_model_column(self.model_file, colno_name)
            )  # Here the already_added_file_names will have no duplicates
            for file_path in file_paths_ok:
                file_name = os_path.splitext(os_path.basename(file_path))[0]
                if file_name in already_added_file_names:
                    occurrence = 2
                    while f"{file_name} ({occurrence})" in already_added_file_names:
                        occurrence += 1
                    file_name = f"{file_name} ({occurrence})"
                already_added_file_names.append(file_name)
                rowno = self.model_file.rowCount()
                self.model_file.set_row_str(rowno, (file_name, file_path))
                self.model_file.data_updated.emit()

        if file_paths_dup or file_paths_unsupported or file_paths_empty:
            model_err_files = Ng_Model(main=self)
            model_err_files.setHorizontalHeaderLabels(("Error Type", "File Path"))
            for reason, file_paths in (
                ("Duplicate file", file_paths_dup),
                ("Unsupported file type", file_paths_unsupported),
                ("Empty file", file_paths_empty),
            ):
                for file_path in file_paths:
                    model_err_files.insertRow(
                        model_err_files.rowCount(),
                        (QStandardItem(reason), QStandardItem(file_path)),
                    )
            tableview_err_files = Ng_TableView(main=self, model=model_err_files, has_vertical_header=False)

            dialog = Ng_Dialog_Table(
                self,
                title="Error Adding Files",
                text="Failed to add the following files.",
                width=300,
                height=200,
                resizable=True,
                tableview=tableview_err_files,
            )
            dialog.exec()

    def menubar_file_open_folder(self):
        # TODO: Currently only include files of supported types, should include
        #  all files, and popup error for unsupported files
        folder_dialog = QFileDialog(self)
        # TODO remove default directory before releasing
        folder_path = folder_dialog.getExistingDirectory(
            caption="Open Folder",
            dir='directory="/home/tan/docx/corpus/YuHua-parallel-corpus-zh-en/02aligned/standalone/',
        )
        if not folder_path:
            return

        file_paths_to_add = []
        for extension in SCAIO.SUPPORTED_EXTENSIONS:
            file_paths_to_add.extend(glob.glob(os_path.join(folder_path, f"*.{extension}")))
        self.add_file_paths(file_paths_to_add)

    def menubar_file_open_file(self):
        file_dialog = QFileDialog(self)
        file_paths_to_add, _ = file_dialog.getOpenFileNames(
            parent=None,
            caption="Open Files",
            dir="/home/tan/docx/corpus/YuHua-parallel-corpus-zh-en/02aligned/standalone/",
            # TODO remove this before releasing
            filter="Text files (*.txt);;Docx files (*.docx);;Odt files (*.odt);;All files (*.*)",
        )
        if not file_paths_to_add:
            return
        self.add_file_paths(file_paths_to_add)

    def menubar_file_restart(self):
        self.close()
        command = [sys.executable, "-m", "neosca_gui"]
        subprocess.call(command, env=os.environ.copy(), close_fds=False)


def main():
    ui_scaling = DEFAULT_INTERFACE_SCALING
    os.environ["QT_SCALE_FACTOR"] = re.sub(r"([0-9]{2})%$", r".\1", ui_scaling)
    ng_app = QApplication(sys.argv)
    ng_window = Ng_Main()
    ng_window.showMaximized()
    sys.exit(ng_app.exec())
