#!/usr/bin/env python3

import json
import os.path as os_path
import re
from enum import Enum
from typing import Dict, Iterable, List, Literal, Optional, Tuple, Union

from PySide6.QtCore import (
    QDir,
    QElapsedTimer,
    QModelIndex,
    QPersistentModelIndex,
    QPoint,
    QTime,
    QTimer,
    Signal,
)
from PySide6.QtGui import (
    QBrush,
    QColor,
    QFocusEvent,
    QPainter,
    QPalette,
    QStandardItem,
    QStandardItemModel,
    Qt,
    QTextBlockFormat,
    QTextCursor,
)
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QComboBox,
    QCompleter,
    QDialog,
    QFileDialog,
    QFileSystemModel,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QStyledItemDelegate,
    QStyleOptionViewItem,
    QTableView,
    QTextEdit,
    QWidget,
)

from neosca_gui import DESKTOP_PATH, NEOSCA_HOME
from neosca_gui.ng_about import __title__
from neosca_gui.ng_qss import Ng_QSS
from neosca_gui.ng_settings.ng_settings import Ng_Settings
from neosca_gui.ng_settings.ng_settings_default import available_export_types
from neosca_gui.ng_singleton import QSingleton


class Ng_Model(QStandardItemModel):
    data_cleared = Signal()
    # data_updated: itemChanged && !data_cleared
    data_updated = Signal()
    data_exported = Signal()

    def __init__(self, *args, main, orientation: Literal["hor", "ver"] = "hor", **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.main = main
        self.orientation = orientation

        self.has_been_exported: bool = False
        self.data_exported.connect(lambda: self.set_has_been_exported(True))
        self.data_updated.connect(lambda: self.set_has_been_exported(False))

    def set_item_str(self, rowno: int, colno: int, value: Union[QStandardItem, str]) -> None:
        item = value if isinstance(value, QStandardItem) else QStandardItem(value)
        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.setItem(rowno, colno, item)

    def set_row_str(self, rowno: int, values: Iterable[Union[QStandardItem, str]]) -> None:
        for colno, value in enumerate(values):
            self.set_item_str(rowno, colno, value)

    def set_item_num(self, rowno: int, colno: int, value: Union[QStandardItem, int, float, str]) -> None:
        if isinstance(value, QStandardItem):
            item = value
        else:
            if not isinstance(value, str):
                value = str(value)
            item = QStandardItem(value)
        item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.setItem(rowno, colno, item)

    def set_row_num(self, rowno: int, values: Iterable[Union[QStandardItem, int, float, str]]) -> None:
        for colno, value in enumerate(values):
            self.set_item_num(rowno, colno, value)

    def set_has_been_exported(self, exported: bool) -> None:
        self.has_been_exported = exported

    def _clear_data(self, leave_an_empty_row=True) -> None:
        if self.orientation == "hor":
            self.setRowCount(0)
            if leave_an_empty_row:
                self.setRowCount(1)
        elif self.orientation == "ver":
            self.setColumnCount(0)
            if leave_an_empty_row:
                self.setColumnCount(1)
        self.data_cleared.emit()

    def clear_data(self, confirm=False, leave_an_empty_row=True) -> None:
        """
        Clear data, reserve headers
        """
        if not confirm or self.has_been_exported:
            return self._clear_data(leave_an_empty_row=leave_an_empty_row)

        messagebox = QMessageBox(self.main)
        messagebox.setWindowTitle("Clear Table")
        messagebox.setText("The table has not been exported yet and all the data will be lost. Continue?")
        messagebox.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        messagebox.accepted.connect(lambda: self._clear_data(leave_an_empty_row=leave_an_empty_row))
        messagebox.exec()

    def is_empty(self):
        for row in range(self.rowCount()):
            for column in range(self.columnCount()):
                item = self.item(row, column)
                if item is not None and item.text() != "":
                    return False
        return True

    def remove_single_empty_row(self) -> None:
        if self.rowCount() == 1 and self.item(0, 0) is None:
            self.setRowCount(0)

    # https://stackoverflow.com/questions/75038194/qt6-how-to-disable-selection-for-empty-cells-in-qtableview
    def flags(self, index) -> Qt.ItemFlag:
        if index.data() is None:
            return Qt.ItemFlag.NoItemFlags
        return super().flags(index)


class Ng_Delegate_SCA(QStyledItemDelegate):
    def __init__(self, parent=None, qss: str = ""):
        super().__init__(parent)
        if (
            triangle_rgb := Ng_QSS.get_value(qss, "QHeaderView::section:vertical", "background-color")
        ) is not None:
            self.triangle_rgb = triangle_rgb
        else:
            self.triangle_rgb = "#000000"

        self.pos_dialog_mappings: Dict[Tuple[int, int], Ng_Dialog_TextEdit_SCA_Matched_Subtrees] = {}

    @staticmethod
    def is_index_clickable(index) -> bool:
        data_in_user_role = index.data(Qt.ItemDataRole.UserRole)
        return data_in_user_role is not None and data_in_user_role

    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index: QModelIndex):
        super().paint(painter, option, index)
        if self.is_index_clickable(index):
            painter.save()
            painter.setBrush(QBrush(QColor.fromString(self.triangle_rgb)))
            triangle_leg_length = option.rect.height() * Ng_Settings.value(
                "Appearance/triangle-height-ratio", type=float
            )
            painter.drawPolygon(
                (
                    QPoint(option.rect.x() + triangle_leg_length, option.rect.y()),
                    QPoint(option.rect.x(), option.rect.y()),
                    QPoint(option.rect.x(), option.rect.y() + triangle_leg_length),
                )
            )
            painter.restore()

    def createEditor(self, parent, option, index):
        if not self.is_index_clickable(index):
            return None
        pos = (index.row(), index.column())
        if pos in self.pos_dialog_mappings:
            self.pos_dialog_mappings[pos].activateWindow()
        else:
            dialog = Ng_Dialog_TextEdit_SCA_Matched_Subtrees(parent, index=index)
            self.pos_dialog_mappings[pos] = dialog
            dialog.finished.connect(lambda: self.pos_dialog_mappings.pop(pos))
            dialog.show()


class Ng_TableView(QTableView):
    def __init__(
        self,
        *args,
        main,
        model: Ng_Model,
        has_horizontal_header: bool = True,
        has_vertical_header: bool = True,
        **kwargs,
    ) -> None:
        super().__init__(*args, **kwargs)
        self.main = main
        self.setModel(model)
        self.model_: Ng_Model = model
        self.model_.data_cleared.connect(self.on_data_cleared)
        self.model_.data_updated.connect(self.on_data_updated)
        self.has_horizontal_header = has_horizontal_header
        self.has_vertical_header = has_vertical_header

        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.horizontalHeader().setHighlightSections(False)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.verticalHeader().setHighlightSections(False)
        self.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)

        if self.model_.is_empty():
            self.setEnabled(False)

    def on_data_cleared(self) -> None:
        self.horizontalHeader().resizeSections()
        self.setEnabled(False)

    def on_data_updated(self) -> None:
        # TODO: only need to enable at the first time
        if not self.isEnabled():
            self.setEnabled(True)
        self.scrollToBottom()

    # Override to specify the return type
    def model(self) -> Ng_Model:
        return self.model_

    def set_openpyxl_horizontal_header_alignment(self, cell) -> None:
        from openpyxl.styles import Alignment

        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def set_openpyxl_vertical_header_alignment(self, cell) -> None:
        from openpyxl.styles import Alignment

        if self.has_vertical_header:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def set_openpyxl_cell_alignment(self, cell, item: QStandardItem) -> None:
        # https://doc.qt.io/qtforpython-6/PySide6/QtCore/Qt.html#PySide6.QtCore.PySide6.QtCore.Qt.AlignmentFlag
        # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.alignment.html
        # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L887
        # Qt
        #  Horizontal: Qt.AlignLeft, Qt.AlignRight,  Qt.AlignHCenter, Qt.AlignJustify
        #  Vertical:   Qt.AlignTop,  Qt.AlignBottom, Qt.AlignVCenter, Qt.AlignBaseline
        # OpenPyXL
        #  Horizontal: justify, center, distributed, left, right, fill, general, centerContinuous
        #  Vertical:   justify, center, distributed, top,  bottom

        from openpyxl.styles import Alignment

        alignment_item: Qt.AlignmentFlag = item.textAlignment()

        # Horizontal
        if alignment_item & Qt.AlignmentFlag.AlignLeft:
            alignment_cell_horizontal = "left"
        elif alignment_item & Qt.AlignmentFlag.AlignRight:
            alignment_cell_horizontal = "right"
        elif alignment_item & Qt.AlignmentFlag.AlignHCenter:
            alignment_cell_horizontal = "center"
        elif alignment_item & Qt.AlignmentFlag.AlignJustify:
            alignment_cell_horizontal = "justify"
        else:
            alignment_cell_horizontal = "left"

        # Vertical
        if Qt.AlignmentFlag.AlignTop in alignment_item:
            alignment_cell_vertical = "top"
        elif Qt.AlignmentFlag.AlignBottom in alignment_item:
            alignment_cell_vertical = "bottom"
        elif Qt.AlignmentFlag.AlignVCenter in alignment_item:
            alignment_cell_vertical = "center"
        # > Wordless: Not sure
        elif Qt.AlignmentFlag.AlignBaseline in alignment_item:
            alignment_cell_vertical = "justify"
        else:
            alignment_cell_vertical = "center"

        cell.alignment = Alignment(
            horizontal=alignment_cell_horizontal, vertical=alignment_cell_vertical, wrap_text=True
        )

    def export_table(self, filename: str = "") -> None:
        file_path, file_type = QFileDialog.getSaveFileName(
            parent=None,
            caption="Export Table",
            dir=str(DESKTOP_PATH / filename),
            filter=";;".join(available_export_types),
            selectedFilter=Ng_Settings.value("Export/default-type"),
        )
        if not file_path:
            return

        model: Ng_Model = self.model()
        col_count = model.columnCount()
        row_count = model.rowCount()
        try:
            if ".xlsx" in file_type:
                # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L701C1-L716C54
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                from openpyxl.utils import get_column_letter

                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet_cell = worksheet.cell

                rowno_cell_offset = 2 if self.has_horizontal_header else 1
                colno_cell_offset = 2 if self.has_vertical_header else 1

                # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L628C3-L629C82
                dpi_horizontal = QApplication.primaryScreen().logicalDotsPerInchX()
                dpi_vertical = QApplication.primaryScreen().logicalDotsPerInchY()

                # 1. Horizontal header text and alignment
                if self.has_horizontal_header:
                    for colno_cell, colno_item in enumerate(range(col_count)):
                        cell = worksheet_cell(1, colno_cell_offset + colno_cell)
                        cell.value = model.horizontalHeaderItem(colno_item).text()
                        self.set_openpyxl_horizontal_header_alignment(cell)
                # 2. Vertical header text and alignment
                if self.has_vertical_header:
                    for rowno_cell, rowno_item in enumerate(range(row_count)):
                        cell = worksheet_cell(rowno_cell_offset + rowno_cell, 1)
                        cell.value = model.verticalHeaderItem(rowno_item).text()
                        self.set_openpyxl_vertical_header_alignment(cell)

                # 3. Both header background and font
                # 3.0.1 Get header background
                horizon_bacolor: Optional[str] = Ng_QSS.get_value(
                    self.main.styleSheet(), "QHeaderView::section:horizontal", "background-color"
                )
                vertical_bacolor: Optional[str] = Ng_QSS.get_value(
                    self.main.styleSheet(), "QHeaderView::section:vertical", "background-color"
                )
                # 3.0.2 Get header font, currently only consider color and boldness
                #  https://www.codespeedy.com/change-font-color-of-excel-cells-using-openpyxl-in-python/
                #  https://doc.qt.io/qt-6/stylesheet-reference.html#font-weight
                font_color = Ng_QSS.get_value(self.main.styleSheet(), "QHeaderView::section", "color")
                font_color = font_color.lstrip("#") if font_color is not None else "000"
                font_weight = Ng_QSS.get_value(self.main.styleSheet(), "QHeaderView::section", "font-weight")
                is_bold = (font_weight == "bold") if font_weight is not None else False
                # 3.1 Horizontal header background and font
                if self.has_horizontal_header:
                    # 3.1.1 Horizontal header background
                    #  TODO: Currently all tabpages share the same style sheet and the
                    #   single QSS file is loaded from MainWindow, thus here the
                    #   style sheet is accessed from self. In the future different
                    #   tabs might load their own QSS files, and the style sheet
                    #   should be accessed from the QTabWidget. This is also the
                    #   case for all other "self.styleSheet()" expressions, when
                    #   making this change, remember to edit all of them.
                    if horizon_bacolor is not None:
                        horizon_bacolor = horizon_bacolor.lstrip("#")
                        for colno in range(col_count):
                            cell = worksheet_cell(1, colno_cell_offset + colno)
                            cell.fill = PatternFill(fill_type="solid", fgColor=horizon_bacolor)
                    # 3.1.2 Horizontal header font
                    for colno in range(col_count):
                        cell = worksheet_cell(1, colno_cell_offset + colno)
                        cell.font = Font(color=font_color, bold=is_bold)
                # 3.2 Vertical header background and font
                if self.has_vertical_header:
                    # 3.2.1 Vertial header background
                    if vertical_bacolor is not None:
                        vertical_bacolor = vertical_bacolor.lstrip("#")
                        for rowno in range(row_count):
                            cell = worksheet_cell(rowno_cell_offset + rowno, 1)
                            cell.fill = PatternFill(fill_type="solid", fgColor=vertical_bacolor)
                    # 3.2.2 Vertical header font
                    for rowno in range(row_count):
                        cell = worksheet_cell(rowno_cell_offset + rowno, 1)
                        cell.font = Font(color=font_color, bold=is_bold)

                # 4. Cells
                for rowno in range(row_count):
                    for colno in range(col_count):
                        cell = worksheet_cell(rowno_cell_offset + rowno, colno_cell_offset + colno)
                        item = model.item(rowno, colno)
                        item_value = item.text()
                        try:  # noqa: SIM105
                            item_value = float(item_value)
                        except ValueError:
                            pass
                        cell.value = item_value
                        self.set_openpyxl_cell_alignment(cell, item)
                # 5. Column width
                for colno in range(col_count):
                    # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L729
                    worksheet.column_dimensions[get_column_letter(colno_cell_offset + colno)].width = (
                        self.horizontalHeader().sectionSize(colno) / dpi_horizontal * 13 + 3
                    )

                if self.has_vertical_header:
                    # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L731
                    worksheet.column_dimensions[get_column_letter(1)].width = (
                        self.verticalHeader().width() / dpi_horizontal * 13 + 3
                    )
                # 6. Row height
                worksheet.row_dimensions[1].height = self.horizontalHeader().height() / dpi_vertical * 72
                # 7. Freeze panes
                # https://stackoverflow.com/questions/73837417/freeze-panes-first-two-rows-and-column-with-openpyxl
                # Using "2" in both cases means to always freeze the 1st column
                if self.has_horizontal_header:
                    worksheet.freeze_panes = "B2"
                else:
                    worksheet.freeze_panes = "A2"
                workbook.save(file_path)
            elif ".csv" in file_type or ".tsv" in file_type:
                import csv

                dialect = csv.excel if ".csv" in file_type else csv.excel_tab
                with open(os_path.normpath(file_path), "w", newline="", encoding="utf-8") as fh:
                    csv_writer = csv.writer(fh, dialect=dialect, lineterminator="\n")
                    # Horizontal header
                    hor_header: List[str] = [""]
                    hor_header.extend(model.horizontalHeaderItem(colno).text() for colno in range(col_count))
                    csv_writer.writerow(hor_header)
                    # Vertical header + cells
                    for rowno in range(row_count):
                        row: List[str] = [model.verticalHeaderItem(rowno).text()]
                        row.extend(model.item(rowno, colno).text() for colno in range(col_count))
                        csv_writer.writerow(row)
            QMessageBox.information(
                self, "Success", f"The table has been successfully exported to {file_path}."
            )
        except PermissionError:
            QMessageBox.critical(
                self, "Error Exporting Cells", f"PermissionError: failed to export the table to {file_path}."
            )
        else:
            model.data_exported.emit()

    def export_matches(self) -> None:
        default_export_dir = os_path.join(DESKTOP_PATH, "neosca_sca_matches")
        if not os_path.isdir(default_export_dir):
            os.makedirs(default_export_dir)
            is_default_dir_just_created = True
        else:
            is_default_dir_just_created = False
        export_dir = QFileDialog.getExistingDirectory(caption="Export Matches", dir=str(DESKTOP_PATH))
        # export_dir = Ng_FileDialog.getSaveFolderName(
        #     caption="Export Matches", dir=desktop, selectedFilter="neosca_sca_matches"
        # )
        if not export_dir:
            return
        if export_dir != default_export_dir and is_default_dir_just_created:
            shutil.rmtree(default_export_dir, ignore_errors=True)

        # TODO: verify folder_path is empty, else pop up dialog warning users
        model = self.model()
        for rowno in range(model.rowCount()):
            rowname = model.verticalHeaderItem(rowno).text()
            for colno in range(model.columnCount()):
                index = model.index(rowno, colno)
                if not Ng_Delegate_SCA.is_index_clickable(index):
                    continue
                nested_folder_path = os_path.join(export_dir, rowname)
                os.makedirs(nested_folder_path, exist_ok=True)
                colname = model.horizontalHeaderItem(colno).text()
                file_path = f"{os_path.join(nested_folder_path, colname)}.txt"
                matches = index.data(Qt.ItemDataRole.UserRole)
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(matches))
        QMessageBox.information(self, "Success", f"The table has been successfully exported to {export_dir}.")


class Ng_FileDialog(QFileDialog):
    def __init__(self, *args, **kwargs) -> None:
        return super().__init__(*args, **kwargs)

    # https://github.com/hibtc/madgui/blob/50d59037eab6e59a4510b5a7c4f953ddee4727f7/src/madgui/widget/filedialog.py#L25
    # https://codebrowser.dev/qt5/qtbase/src/widgets/dialogs/qfiledialog.cpp.html#_ZN11QFileDialog14getSaveFileUrlEP7QWidgetRK7QStringRK4QUrlS4_PS2_6QFlagsINS_6OptionEERK11QStringList
    @classmethod
    def getSaveFolderName(
        cls,
        parent: Optional[QWidget] = None,
        caption: str = "",
        dir: str = "",
        filter: str = "",
        selectedFilter: Optional[str] = None,
    ):
        dialog = cls(parent, caption, dir, filter)
        dialog.setOptions(QFileDialog.Option.ShowDirsOnly)
        dialog.setFileMode(QFileDialog.FileMode.Directory)
        dialog.setAcceptMode(QFileDialog.AcceptMode.AcceptSave)
        if selectedFilter is not None and selectedFilter:
            dialog.selectNameFilter(selectedFilter)

        if dialog.exec() != QFileDialog.DialogCode.Accepted:
            return None

        # Returns a list of strings containing the absolute paths of the
        # selected files in the dialog. If no files are selected, or the mode
        # is not ExistingFiles or ExistingFile, selectedFiles() contains the
        # current path in the viewport.
        filename = dialog.selectedFiles()[0]
        return filename


class Ng_Dialog(QDialog):
    class ButtonAlignmentFlag(Enum):
        AlignLeft = 0
        AlignRight = 2

    def __init__(
        self, *args, title: str = "", width: int = 0, height: int = 0, resizable=False, **kwargs
    ) -> None:
        """
        ┌———————————┐
        │           │
        │  content  │
        │           │
        │———————————│
        │  buttons  │
        └———————————┘
        """
        super().__init__(*args, **kwargs)
        # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_dialogs/wl_dialogs.py#L28
        # [Copied code starts here]
        # Dialog size
        if resizable:
            if not width:
                width = self.size().width()

            if not height:
                height = self.size().height()

            self.resize(width, height)
        else:
            if width:
                self.setFixedWidth(width)

            if height:
                self.setFixedHeight(height)
            # Gives the window a thin dialog border on Windows. This style is
            # traditionally used for fixed-size dialogs.
            self.setWindowFlag(Qt.WindowType.MSWindowsFixedSizeDialogHint)
        # [Copied code ends here]
        self.setWindowTitle(title)

        self.content_layout = QGridLayout()
        self.button_layout = QGridLayout()
        self.button_layout.setColumnStretch(1, 1)

        self.grid_layout = QGridLayout()
        self.grid_layout.addLayout(self.content_layout, 0, 0)
        self.grid_layout.addLayout(self.button_layout, 1, 0)
        self.setLayout(self.grid_layout)

    def rowCount(self) -> int:
        return self.content_layout.rowCount()

    def columnCount(self) -> int:
        return self.content_layout.columnCount()

    def addWidget(self, *args, **kwargs) -> None:
        self.content_layout.addWidget(*args, **kwargs)

    def addButtons(self, *buttons, alignment: ButtonAlignmentFlag) -> None:
        layout = QGridLayout()
        for colno, button in enumerate(buttons):
            layout.addWidget(button, 0, colno)
        self.button_layout.addLayout(layout, 0, alignment.value)

    def setColumnStretch(self, column: int, strech: int) -> None:
        self.content_layout.setColumnStretch(column, strech)

    def setRowStretch(self, row: int, strech: int) -> None:
        self.content_layout.setRowStretch(row, strech)


class Ng_Dialog_Processing_With_Elapsed_Time(Ng_Dialog):
    started = Signal()
    # Use this to get the place holder, e.g. 0:00:00
    time_format_re = re.compile(r"[^:]")

    def __init__(
        self,
        *args,
        title: str = "Please Wait",
        width: int = 500,
        height: int = 0,
        time_format: str = "h:mm:ss",
        interval: int = 1000,
        **kwargs,
    ) -> None:
        super().__init__(*args, title=title, width=width, height=height, resizable=False, **kwargs)
        self.time_format = time_format
        self.interval = interval
        self.elapsedtimer = QElapsedTimer()
        self.timer = QTimer()

        # TODO: this label should be exposed
        self.label_status = QLabel("Processing...")
        self.text_time_elapsed_zero = f"Elapsed time: {self.time_format_re.sub('0', time_format)}"
        self.label_time_elapsed = QLabel(self.text_time_elapsed_zero)
        self.label_please_wait = QLabel(
            "Please be patient. The wait time can range from a few seconds to several minutes."
        )
        self.label_please_wait.setWordWrap(True)

        self.addWidget(self.label_status, 0, 0)
        self.addWidget(self.label_time_elapsed, 0, 1, alignment=Qt.AlignmentFlag.AlignRight)
        self.addWidget(self.label_please_wait, 1, 0, 1, 2)

        self.setWindowFlags(self.windowFlags() | Qt.WindowType.FramelessWindowHint)

        # Bind
        self.timer.timeout.connect(self.update_time_elapsed)
        self.started.connect(self.elapsedtimer.start)
        # If the timer is already running, it will be stopped and restarted.
        self.started.connect(lambda: self.timer.start(self.interval))
        # Either 'accepted' or 'rejected', although 'rejected' is disabled (see rejected below)
        self.finished.connect(self.reset_time_elapsed)
        self.finished.connect(self.timer.stop)

    def reset_time_elapsed(self) -> None:
        self.label_time_elapsed.setText(self.text_time_elapsed_zero)

    def update_time_elapsed(self) -> None:
        time_elapsed: int = self.elapsedtimer.elapsed()
        qtime: QTime = QTime.fromMSecsSinceStartOfDay(time_elapsed)
        self.label_time_elapsed.setText(f"Elapsed time: {qtime.toString(self.time_format)}")

    # Override
    def reject(self) -> None:
        pass

    # Override
    def show(self) -> None:
        self.started.emit()
        return super().show()

    # Override
    def open(self) -> None:
        self.started.emit()
        return super().open()

    # Override
    def exec(self) -> int:
        self.started.emit()
        return super().exec()


class Ng_Dialog_TextEdit(Ng_Dialog):
    def __init__(self, *args, title: str = "", text: str = "", **kwargs) -> None:
        super().__init__(*args, title=title, resizable=True, **kwargs)
        self.textedit = QTextEdit(text)
        self.textedit.setReadOnly(True)
        # https://stackoverflow.com/questions/74852753/indent-while-line-wrap-on-qtextedit-with-pyside6-pyqt6
        indentation: int = self.fontMetrics().horizontalAdvance("abcd")
        self.fmt_textedit = QTextBlockFormat()
        self.fmt_textedit.setLeftMargin(indentation)
        self.fmt_textedit.setTextIndent(-indentation)

        self.button_copy = QPushButton("Copy")
        self.button_copy.clicked.connect(self.textedit.selectAll)
        self.button_copy.clicked.connect(self.textedit.copy)

        self.button_close = QPushButton("Close")
        self.button_close.clicked.connect(self.reject)

        self.addButtons(self.button_copy, alignment=Ng_Dialog.ButtonAlignmentFlag.AlignLeft)
        self.addButtons(self.button_close, alignment=Ng_Dialog.ButtonAlignmentFlag.AlignRight)

    def setText(self, text: str) -> None:
        self.textedit.setText(text)
        cursor = QTextCursor(self.textedit.document())
        cursor.select(QTextCursor.SelectionType.Document)
        cursor.mergeBlockFormat(self.fmt_textedit)

    # Override
    def show(self) -> None:
        # Add self.textedit lastly to allow users add custom widgets above
        self.addWidget(self.textedit, self.rowCount(), 0, 1, self.columnCount())
        return super().show()

    # Override
    def open(self) -> None:
        # Add self.textedit lastly to allow users add custom widgets above
        self.addWidget(self.textedit, self.rowCount(), 0, 1, self.columnCount())
        return super().open()

    # Override
    def exec(self) -> int:
        # Add self.textedit lastly to allow users add custom widgets above
        self.addWidget(self.textedit, self.rowCount(), 0, 1, self.columnCount())
        return super().exec()


class Ng_Dialog_TextEdit_SCA_Matched_Subtrees(Ng_Dialog_TextEdit):
    def __init__(self, *args, index: Union[QModelIndex, QPersistentModelIndex], **kwargs):
        super().__init__(*args, title="Matches", width=500, height=300, **kwargs)
        self.file_name = index.model().verticalHeaderItem(index.row()).text()
        self.sname = index.model().horizontalHeaderItem(index.column()).text()
        self.matched_subtrees: List[str] = index.data(Qt.ItemDataRole.UserRole)
        self.setText("\n".join(self.matched_subtrees))

        self.label_summary = QLabel(
            f'{len(self.matched_subtrees)} occurrences of "{self.sname}" in "{self.file_name}"'
        )
        self.label_summary.setWordWrap(True)
        self.addWidget(self.label_summary)


class Ng_Dialog_TextEdit_Citing(Ng_Dialog_TextEdit):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, title="Citing", **kwargs)
        # citing.json is at the same dir of __file__
        # TODO: need to have a unified way to get project path.
        with open(NEOSCA_HOME / "citing.json", encoding="utf-8") as f:
            self.style_citation_mapping = json.load(f)

        self.label_citing = QLabel(f"If you use {__title__} in your research, please kindly cite as follows.")
        self.label_citing.setWordWrap(True)
        self.setText(next(iter(self.style_citation_mapping.values())))
        self.label_choose_citation_style = QLabel("Choose citation style: ")
        self.combobox_choose_citation_style = QComboBox()
        self.combobox_choose_citation_style.addItems(tuple(self.style_citation_mapping.keys()))
        self.combobox_choose_citation_style.currentTextChanged.connect(
            lambda key: self.setText(self.style_citation_mapping[key])
        )

        self.addWidget(self.label_citing, 0, 0, 1, 2)
        self.addWidget(self.label_choose_citation_style, 1, 0)
        self.addWidget(self.combobox_choose_citation_style, 1, 1)
        self.setColumnStretch(1, 1)


class Ng_Dialog_Table(Ng_Dialog):
    def __init__(
        self,
        *args,
        text: str,
        tableview: Ng_TableView,
        export_filename: str = "",
        **kwargs,
    ) -> None:
        super().__init__(*args, **kwargs)
        self.tableview: Ng_TableView = tableview
        self.content_layout.addWidget(QLabel(text), 0, 0)
        self.content_layout.addWidget(tableview, 1, 0)
        self.export_filename = export_filename

        self.button_ok = QPushButton("OK")
        self.button_ok.clicked.connect(self.accept)
        self.button_export_table = QPushButton("Export table...")
        self.button_export_table.clicked.connect(lambda: self.tableview.export_table(self.export_filename))
        self.button_layout.addWidget(self.button_export_table, 0, 0, alignment=Qt.AlignmentFlag.AlignLeft)
        self.button_layout.addWidget(self.button_ok, 0, 1, alignment=Qt.AlignmentFlag.AlignRight)


# https://github.com/BLKSerene/Wordless/blob/fa743bcc2a366ec7a625edc4ed6cfc355b7cd22e/wordless/wl_widgets/wl_layouts.py#L108
class Ng_ScrollArea(QScrollArea):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWidgetResizable(True)
        self.setBackgroundRole(QPalette.ColorRole.Light)


class Ng_FileSystemModel(QFileSystemModel, metaclass=QSingleton):
    def __init__(self, parent=None):
        super().__init__(parent)
        # > Do not add file watchers to the paths. This reduces overhead when using the
        # > model for simple tasks like line edit completion.
        self.setOption(QFileSystemModel.Option.DontWatchForChanges)
        self.has_set_root = False

    def start_querying(self):
        # > QFileSystemModel will not fetch any files or directories until
        # > setRootPath() is called.
        if not self.has_set_root:
            self.setRootPath(QDir.homePath())
            self.has_set_root = True


class Ng_LineEdit(QLineEdit):
    """This class emits the custom "focused" signal and is specifically used
    in Ng_LineEdit_Path to tell Ng_FileSystemModel to start querying. The
    querying should only start at the first emit and all subsequent emits are
    ignored. We prefer the custom "focused" signal over the built-in
    "textEdited" because it has much less frequent emits."""

    focused = Signal()

    def __init__(self, contents: Optional[str] = None, parent: Optional[QWidget] = None):
        if contents is None:
            super().__init__(parent)
        else:
            super().__init__(contents, parent)

    # Override
    def focusInEvent(self, e: QFocusEvent):
        super().focusInEvent(e)
        self.focused.emit()


class Ng_LineEdit_Path(QWidget):
    # https://stackoverflow.com/a/20796318/20732031
    def __init__(self, parent=None):
        super().__init__(parent)

        filesystem_model = Ng_FileSystemModel()
        completer_lineedit_files = QCompleter()
        completer_lineedit_files.setModel(filesystem_model)
        completer_lineedit_files.setCompletionMode(QCompleter.CompletionMode.InlineCompletion)
        completer_lineedit_files.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self.lineedit = Ng_LineEdit()
        self.lineedit.focused.connect(filesystem_model.start_querying)
        self.lineedit.setCompleter(completer_lineedit_files)
        self.lineedit.setClearButtonEnabled(True)
        button_browse = QPushButton("Browse")

        # Bind
        button_browse.clicked.connect(self.browse_path)

        hlayout = QHBoxLayout()
        hlayout.setContentsMargins(0, 0, 0, 0)
        hlayout.addWidget(self.lineedit)
        hlayout.addWidget(button_browse)
        self.setLayout(hlayout)

    def text(self) -> str:
        return self.lineedit.text()

    def setText(self, text: str) -> None:
        self.lineedit.setText(text)

    def browse_path(self):
        folder_path = QFileDialog.getExistingDirectory(caption="Choose Path")
        if not folder_path:
            return
        self.lineedit.setText(folder_path)

    def setFocus(self) -> None:
        self.lineedit.setFocus()

    def selectAll(self) -> None:
        self.lineedit.selectAll()


class Ng_Combobox_Editable(QComboBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        # https://stackoverflow.com/questions/45393507/pyqt4-avoid-adding-the-items-to-the-qcombobox
        self.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.setEditable(True)
