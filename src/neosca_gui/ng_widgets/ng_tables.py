#!/usr/bin/env python3

import os.path as os_path
from typing import Dict, Generator, Iterable, List, Literal, Optional, Tuple, Union

from PySide6.QtCore import (
    QModelIndex,
    QPoint,
    Signal,
)
from PySide6.QtGui import (
    QBrush,
    QColor,
    QPainter,
    QStandardItem,
    QStandardItemModel,
    Qt,
)
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QFileDialog,
    QHeaderView,
    QMessageBox,
    QStyledItemDelegate,
    QStyleOptionViewItem,
    QTableView,
)

from neosca_gui import DESKTOP_PATH
from neosca_gui.ng_qss import Ng_QSS
from neosca_gui.ng_settings.ng_settings import Ng_Settings
from neosca_gui.ng_settings.ng_settings_default import available_export_types
from neosca_gui.ng_widgets.ng_dialogs import Ng_Dialog_TextEdit_SCA_Matched_Subtrees
from neosca_gui.ng_widgets.ng_widgets import Ng_MessageBox_Confirm


class Ng_StandardItemModel(QStandardItemModel):
    data_cleared = Signal()
    # data_updated: itemChanged && !data_cleared
    data_updated = Signal()
    data_exported = Signal()

    def __init__(self, *args, main, orientation: Literal["hor", "ver"] = "hor", **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.main = main
        self.orientation = orientation

        self.has_been_exported: bool = False
        self.data_exported.connect(lambda: self.set_has_been_exported(True))
        self.data_updated.connect(lambda: self.set_has_been_exported(False))

    def set_item_str(self, rowno: int, colno: int, value: Union[QStandardItem, str]) -> None:
        item = value if isinstance(value, QStandardItem) else QStandardItem(value)
        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.setItem(rowno, colno, item)

    def set_row_str(self, rowno: int, values: Iterable[Union[QStandardItem, str]]) -> None:
        for colno, value in enumerate(values):
            self.set_item_str(rowno, colno, value)

    def set_item_num(self, rowno: int, colno: int, value: Union[QStandardItem, int, float, str]) -> None:
        if isinstance(value, QStandardItem):
            item = value
        else:
            if not isinstance(value, str):
                value = str(value)
            item = QStandardItem(value)
        item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.setItem(rowno, colno, item)

    def set_row_num(self, rowno: int, values: Iterable[Union[QStandardItem, int, float, str]]) -> None:
        for colno, value in enumerate(values):
            self.set_item_num(rowno, colno, value)

    def set_has_been_exported(self, exported: bool) -> None:
        self.has_been_exported = exported

    def _clear_data(self, leave_an_empty_row=True) -> None:
        if self.orientation == "hor":
            self.setRowCount(0)
            if leave_an_empty_row:
                self.setRowCount(1)
        elif self.orientation == "ver":
            self.setColumnCount(0)
            if leave_an_empty_row:
                self.setColumnCount(1)
        self.data_cleared.emit()

    def clear_data(self, confirm=False, leave_an_empty_row=True) -> None:
        """
        Clear data, reserve headers
        """
        if not confirm or self.has_been_exported:
            return self._clear_data(leave_an_empty_row=leave_an_empty_row)
        else:
            messagebox = Ng_MessageBox_Confirm(
                self.main,
                "Clear Talbe",
                "The table has not been exported yet and all the data will be lost. Continue?",
            )
            if messagebox.exec():
                self._clear_data(leave_an_empty_row=leave_an_empty_row)

    def is_empty(self):
        for row in range(self.rowCount()):
            for column in range(self.columnCount()):
                item = self.item(row, column)
                if item is not None and item.text() != "":
                    return False
        return True

    def remove_single_empty_row(self) -> None:
        if self.rowCount() == 1 and self.item(0, 0) is None:
            self.setRowCount(0)

    def remove_model_rows(self, *rownos: int) -> None:
        if not rownos:
            # https://doc.qt.io/qtforpython-6/PySide6/QtGui/QStandardItemModel.html#PySide6.QtGui.PySide6.QtGui.QStandardItemModel.setRowCount
            self.setRowCount(0)
        else:
            for rowno in rownos:
                self.takeRow(rowno)

    # Type hint for generator: https://docs.python.org/3.12/library/typing.html#typing.Generator
    def yield_model_column(self, colno: int) -> Generator[str, None, None]:
        items = (self.item(rowno, colno) for rowno in range(self.rowCount()))
        return (item.text() for item in items if item is not None)

    # https://stackoverflow.com/questions/75038194/qt6-how-to-disable-selection-for-empty-cells-in-qtableview
    def flags(self, index) -> Qt.ItemFlag:
        if index.data() is None:
            return Qt.ItemFlag.NoItemFlags
        return super().flags(index)


class Ng_Delegate_SCA(QStyledItemDelegate):
    def __init__(self, parent=None, qss: str = ""):
        super().__init__(parent)
        if (
            triangle_rgb := Ng_QSS.get_value(qss, "QHeaderView::section:vertical", "background-color")
        ) is not None:
            self.triangle_rgb = triangle_rgb
        else:
            self.triangle_rgb = "#000000"

        self.pos_dialog_mappings: Dict[Tuple[int, int], Ng_Dialog_TextEdit_SCA_Matched_Subtrees] = {}

    @staticmethod
    def is_index_clickable(index) -> bool:
        data_in_user_role = index.data(Qt.ItemDataRole.UserRole)
        return data_in_user_role is not None and data_in_user_role

    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index: QModelIndex):
        super().paint(painter, option, index)
        if self.is_index_clickable(index):
            painter.save()
            painter.setBrush(QBrush(QColor.fromString(self.triangle_rgb)))
            triangle_leg_length = option.rect.height() * Ng_Settings.value(
                "Appearance/triangle-height-ratio", type=float
            )
            painter.drawPolygon(
                (
                    QPoint(option.rect.x() + triangle_leg_length, option.rect.y()),
                    QPoint(option.rect.x(), option.rect.y()),
                    QPoint(option.rect.x(), option.rect.y() + triangle_leg_length),
                )
            )
            painter.restore()

    def createEditor(self, parent, option, index):
        if not self.is_index_clickable(index):
            return None
        pos = (index.row(), index.column())
        if pos in self.pos_dialog_mappings:
            self.pos_dialog_mappings[pos].activateWindow()
        else:
            dialog = Ng_Dialog_TextEdit_SCA_Matched_Subtrees(parent, index=index)
            self.pos_dialog_mappings[pos] = dialog
            dialog.finished.connect(lambda: self.pos_dialog_mappings.pop(pos))
            dialog.show()


class Ng_TableView(QTableView):
    def __init__(
        self,
        *args,
        main,
        model: Ng_StandardItemModel,
        has_horizontal_header: bool = True,
        has_vertical_header: bool = True,
        **kwargs,
    ) -> None:
        super().__init__(*args, **kwargs)
        self.main = main
        self.setModel(model)
        self.model_: Ng_StandardItemModel = model
        self.model_.data_cleared.connect(self.on_data_cleared)
        self.model_.data_updated.connect(self.on_data_updated)
        self.has_horizontal_header = has_horizontal_header
        self.has_vertical_header = has_vertical_header

        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.horizontalHeader().setHighlightSections(False)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.verticalHeader().setHighlightSections(False)
        self.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)

        if self.model_.is_empty():
            self.setEnabled(False)

    def on_data_cleared(self) -> None:
        self.horizontalHeader().resizeSections()
        self.setEnabled(False)

    def on_data_updated(self) -> None:
        # TODO: only need to enable at the first time
        if not self.isEnabled():
            self.setEnabled(True)
        self.scrollToBottom()

    # Override to specify the return type
    def model(self) -> Ng_StandardItemModel:
        return self.model_

    def set_openpyxl_horizontal_header_alignment(self, cell) -> None:
        from openpyxl.styles import Alignment

        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def set_openpyxl_vertical_header_alignment(self, cell) -> None:
        from openpyxl.styles import Alignment

        if self.has_vertical_header:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def set_openpyxl_cell_alignment(self, cell, item: QStandardItem) -> None:
        # https://doc.qt.io/qtforpython-6/PySide6/QtCore/Qt.html#PySide6.QtCore.PySide6.QtCore.Qt.AlignmentFlag
        # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.alignment.html
        # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L887
        # Qt
        #  Horizontal: Qt.AlignLeft, Qt.AlignRight,  Qt.AlignHCenter, Qt.AlignJustify
        #  Vertical:   Qt.AlignTop,  Qt.AlignBottom, Qt.AlignVCenter, Qt.AlignBaseline
        # OpenPyXL
        #  Horizontal: justify, center, distributed, left, right, fill, general, centerContinuous
        #  Vertical:   justify, center, distributed, top,  bottom

        from openpyxl.styles import Alignment

        alignment_item: Qt.AlignmentFlag = item.textAlignment()

        # Horizontal
        if alignment_item & Qt.AlignmentFlag.AlignLeft:
            alignment_cell_horizontal = "left"
        elif alignment_item & Qt.AlignmentFlag.AlignRight:
            alignment_cell_horizontal = "right"
        elif alignment_item & Qt.AlignmentFlag.AlignHCenter:
            alignment_cell_horizontal = "center"
        elif alignment_item & Qt.AlignmentFlag.AlignJustify:
            alignment_cell_horizontal = "justify"
        else:
            alignment_cell_horizontal = "left"

        # Vertical
        if Qt.AlignmentFlag.AlignTop in alignment_item:
            alignment_cell_vertical = "top"
        elif Qt.AlignmentFlag.AlignBottom in alignment_item:
            alignment_cell_vertical = "bottom"
        elif Qt.AlignmentFlag.AlignVCenter in alignment_item:
            alignment_cell_vertical = "center"
        # > Wordless: Not sure
        elif Qt.AlignmentFlag.AlignBaseline in alignment_item:
            alignment_cell_vertical = "justify"
        else:
            alignment_cell_vertical = "center"

        cell.alignment = Alignment(
            horizontal=alignment_cell_horizontal, vertical=alignment_cell_vertical, wrap_text=True
        )

    def export_table(self, filename: str = "") -> None:
        file_path, file_type = QFileDialog.getSaveFileName(
            parent=None,
            caption="Export Table",
            dir=str(DESKTOP_PATH / filename),
            filter=";;".join(available_export_types),
            selectedFilter=Ng_Settings.value("Export/default-type"),
        )
        if not file_path:
            return

        model: Ng_StandardItemModel = self.model()
        col_count = model.columnCount()
        row_count = model.rowCount()
        try:
            if ".xlsx" in file_type:
                # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L701C1-L716C54
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                from openpyxl.utils import get_column_letter

                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet_cell = worksheet.cell

                rowno_cell_offset = 2 if self.has_horizontal_header else 1
                colno_cell_offset = 2 if self.has_vertical_header else 1

                # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L628C3-L629C82
                dpi_horizontal = QApplication.primaryScreen().logicalDotsPerInchX()
                dpi_vertical = QApplication.primaryScreen().logicalDotsPerInchY()

                font_size = Ng_Settings.value("Appearance/font-size", type=int)

                # 1. Horizontal header text and alignment
                if self.has_horizontal_header:
                    for colno_cell, colno_item in enumerate(range(col_count)):
                        cell = worksheet_cell(1, colno_cell_offset + colno_cell)
                        cell.value = model.horizontalHeaderItem(colno_item).text()
                        self.set_openpyxl_horizontal_header_alignment(cell)
                # 2. Vertical header text and alignment
                if self.has_vertical_header:
                    for rowno_cell, rowno_item in enumerate(range(row_count)):
                        cell = worksheet_cell(rowno_cell_offset + rowno_cell, 1)
                        cell.value = model.verticalHeaderItem(rowno_item).text()
                        self.set_openpyxl_vertical_header_alignment(cell)

                # 3. Both header background and font
                # 3.0.1 Get header background
                horizon_bacolor: Optional[str] = Ng_QSS.get_value(
                    self.main.styleSheet(), "QHeaderView::section:horizontal", "background-color"
                )
                vertical_bacolor: Optional[str] = Ng_QSS.get_value(
                    self.main.styleSheet(), "QHeaderView::section:vertical", "background-color"
                )
                # 3.0.2 Get header font, currently only consider color and boldness
                #  https://www.codespeedy.com/change-font-color-of-excel-cells-using-openpyxl-in-python/
                #  https://doc.qt.io/qt-6/stylesheet-reference.html#font-weight
                header_font_color = Ng_QSS.get_value(self.main.styleSheet(), "QHeaderView::section", "color")
                header_font_color = header_font_color.lstrip("#") if header_font_color is not None else "000"
                header_font_weight = Ng_QSS.get_value(
                    self.main.styleSheet(), "QHeaderView::section", "font-weight"
                )
                header_is_bold = (header_font_weight == "bold") if header_font_weight is not None else False
                # 3.1 Horizontal header background and font
                if self.has_horizontal_header:
                    # 3.1.1 Horizontal header background
                    #  TODO: Currently all tabpages share the same style sheet and the
                    #   single QSS file is loaded from MainWindow, thus here the
                    #   style sheet is accessed from self. In the future different
                    #   tabs might load their own QSS files, and the style sheet
                    #   should be accessed from the QTabWidget. This is also the
                    #   case for all other "self.styleSheet()" expressions, when
                    #   making this change, remember to edit all of them.
                    if horizon_bacolor is not None:
                        horizon_bacolor = horizon_bacolor.lstrip("#")
                        for colno in range(col_count):
                            cell = worksheet_cell(1, colno_cell_offset + colno)
                            cell.fill = PatternFill(fill_type="solid", fgColor=horizon_bacolor)
                    # 3.1.2 Horizontal header font
                    for colno in range(col_count):
                        cell = worksheet_cell(1, colno_cell_offset + colno)
                        cell.font = Font(
                            color=header_font_color,
                            bold=header_is_bold,
                            size=font_size,
                        )
                # 3.2 Vertical header background and font
                if self.has_vertical_header:
                    # 3.2.1 Vertial header background
                    if vertical_bacolor is not None:
                        vertical_bacolor = vertical_bacolor.lstrip("#")
                        for rowno in range(row_count):
                            cell = worksheet_cell(rowno_cell_offset + rowno, 1)
                            cell.fill = PatternFill(fill_type="solid", fgColor=vertical_bacolor)
                    # 3.2.2 Vertical header font
                    for rowno in range(row_count):
                        cell = worksheet_cell(rowno_cell_offset + rowno, 1)
                        cell.font = Font(
                            color=header_font_color,
                            bold=header_is_bold,
                            size=font_size,
                        )

                # 4. Cells
                for rowno in range(row_count):
                    for colno in range(col_count):
                        cell = worksheet_cell(rowno_cell_offset + rowno, colno_cell_offset + colno)
                        item = model.item(rowno, colno)
                        item_value = item.text()
                        try:  # noqa: SIM105
                            item_value = float(item_value)
                        except ValueError:
                            pass
                        cell.value = item_value
                        self.set_openpyxl_cell_alignment(cell, item)
                        cell.font = Font(size=font_size)
                # 5. Column width
                for colno in range(col_count):
                    # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L729
                    worksheet.column_dimensions[get_column_letter(colno_cell_offset + colno)].width = (
                        self.horizontalHeader().sectionSize(colno) / dpi_horizontal * 13 + 3
                    )

                if self.has_vertical_header:
                    # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L731
                    worksheet.column_dimensions[get_column_letter(1)].width = (
                        self.verticalHeader().width() / dpi_horizontal * 13 + 3
                    )
                # 6. Row height
                worksheet.row_dimensions[1].height = self.horizontalHeader().height() / dpi_vertical * 72
                for i, _ in enumerate(worksheet.rows):
                    worksheet.row_dimensions[2 + i].height = (
                        self.verticalHeader().sectionSize(0) / dpi_vertical * 72
                    )

                # 7. Freeze panes
                # https://stackoverflow.com/questions/73837417/freeze-panes-first-two-rows-and-column-with-openpyxl
                # Using "2" in both cases means to always freeze the 1st column
                if self.has_horizontal_header:
                    worksheet.freeze_panes = "B2"
                else:
                    worksheet.freeze_panes = "A2"
                workbook.save(file_path)
            elif ".csv" in file_type or ".tsv" in file_type:
                import csv

                dialect = csv.excel if ".csv" in file_type else csv.excel_tab
                with open(os_path.normpath(file_path), "w", newline="", encoding="utf-8") as fh:
                    csv_writer = csv.writer(fh, dialect=dialect, lineterminator="\n")
                    # Horizontal header
                    hor_header: List[str] = [""]
                    hor_header.extend(model.horizontalHeaderItem(colno).text() for colno in range(col_count))
                    csv_writer.writerow(hor_header)
                    # Vertical header + cells
                    for rowno in range(row_count):
                        row: List[str] = [model.verticalHeaderItem(rowno).text()]
                        row.extend(model.item(rowno, colno).text() for colno in range(col_count))
                        csv_writer.writerow(row)
            QMessageBox.information(
                self, "Success", f"The table has been successfully exported to {file_path}."
            )
        except PermissionError:
            QMessageBox.critical(
                self, "Error Exporting Cells", f"PermissionError: failed to export the table to {file_path}."
            )
        else:
            model.data_exported.emit()

    def export_matches(self) -> None:
        file_path, file_type = QFileDialog.getSaveFileName(
            parent=None,
            caption="Export Table",
            dir=str(DESKTOP_PATH / "neosca_sca_matches.xlsx"),
            filter=";;".join(available_export_types),
            selectedFilter=Ng_Settings.value("Export/default-type"),
        )
        if not file_path:
            return

        model = self.model()
        col_count = model.columnCount()
        row_count = model.rowCount()
        try:
            if ".xlsx" in file_type:
                import openpyxl
                from openpyxl.styles import Alignment, Font
                from openpyxl.utils import get_column_letter

                workbook = openpyxl.Workbook()
                for sheetname in workbook.sheetnames:
                    workbook.remove(workbook.get_sheet_by_name(sheetname))

                # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L628C3-L629C82
                dpi_horizontal = QApplication.primaryScreen().logicalDotsPerInchX()
                dpi_vertical = QApplication.primaryScreen().logicalDotsPerInchY()

                font_size = Ng_Settings.value("Appearance/font-size", type=int)

                # Number of matches is much more than that of files. There are
                # 2000 matches across different structures for a 19KB test
                # file. Excel has the max row limit 1,048,576 on a worksheet,
                # which can hold matches of ~524 19KB test files. Here split
                # the matches into different worksheets in case users' corpus
                # is larger than that.
                # https://support.microsoft.com/en-us/office/excel-specifications-and-limits-1672b34d-7043-467e-8e27-269d656771c3
                horizon_header_colwith = 0
                for colno in range(col_count):
                    structure = model.horizontalHeaderItem(colno).text()
                    for rowno in range(row_count):
                        index = model.index(rowno, colno)
                        # TODO: this func is meant to be generic, write an
                        # abstract class of Ng_Delegate_SCA and Ng_Delegate_LCA
                        # (coming), and use the abstract class'
                        # is_index_clickable method
                        if not Ng_Delegate_SCA.is_index_clickable(index):
                            continue

                        filename = model.verticalHeaderItem(rowno).text()
                        if filename not in workbook.sheetnames:
                            ws = workbook.create_sheet(filename)
                            ws_max_row_offset = 0
                        else:
                            ws = workbook[filename]
                            ws_max_row_offset = 1

                        ws_cell = ws.cell
                        ws_start_rowno = ws.max_row + ws_max_row_offset
                        matches: List[str] = index.data(Qt.ItemDataRole.UserRole)
                        for i, match in enumerate(matches):
                            cell_structure = ws_cell(ws_start_rowno + i, 1)
                            cell_structure.value = structure
                            self.set_openpyxl_horizontal_header_alignment(cell_structure)

                            cell_match = ws_cell(ws_start_rowno + i, 2)
                            cell_match.value = match
                            cell_match.alignment = Alignment(horizontal="left", vertical="center")
                            cell_match.font = Font(size=font_size)

                    horizon_header_colwith = max(
                        horizon_header_colwith, self.horizontalHeader().sectionSize(colno)
                    )

                for ws in workbook.worksheets:
                    # Column width
                    ws.column_dimensions[get_column_letter(1)].width = (
                        horizon_header_colwith / dpi_horizontal * 13 + 3
                    )
                    # Row height
                    for i, _ in enumerate(ws.rows):
                        ws.row_dimensions[1 + i].height = self.horizontalHeader().height() / dpi_vertical * 72

                # Header background color
                # horizon_bacolor: Optional[str] = Ng_QSS.get_value(
                #     self.main.styleSheet(), "QHeaderView::section:horizontal", "background-color"
                # )
                # if horizon_bacolor is not None:
                #     horizon_bacolor = horizon_bacolor.lstrip("#")
                #     for ws in workbook.worksheets:
                #         for cell in ws[get_column_letter(1)]:
                #             cell.fill = PatternFill(fill_type="solid", fgColor=horizon_bacolor)
                vertical_bacolor: Optional[str] = Ng_QSS.get_value(
                    self.main.styleSheet(), "QHeaderView::section:vertical", "background-color"
                )
                if vertical_bacolor is not None:
                    vertical_bacolor = vertical_bacolor.lstrip("#")
                    for ws in workbook.worksheets:
                        # https://openpyxl.readthedocs.io/en/stable/worksheet_properties.html#available-properties-for-worksheets
                        ws.sheet_properties.tabColor = vertical_bacolor
                        ws.sheet_view.showGridLines = False
                # Header font
                for ws in workbook.worksheets:
                    for cell in ws[get_column_letter(1)]:
                        cell.font = Font(size=Ng_Settings.value("Appearance/font-size", type=int))

                # Freeze panes
                for ws in workbook.worksheets:
                    ws.freeze_panes = "B1"

                workbook.save(file_path)
            elif ".csv" in file_type or ".tsv" in file_type:
                import csv

                dialect = csv.excel if ".csv" in file_type else csv.excel_tab
                with open(os_path.normpath(file_path), "w", newline="", encoding="utf-8") as fh:
                    csv_writer = csv.writer(fh, dialect=dialect, lineterminator="\n")
                    header: List[str] = ["Filename", "Structure", "Match"]
                    csv_writer.writerow(header)
                    for rowno in range(row_count):
                        filename = model.verticalHeaderItem(rowno).text()
                        for colno in range(col_count):
                            index = model.index(rowno, colno)
                            if not Ng_Delegate_SCA.is_index_clickable(index):
                                continue
                            structure = model.horizontalHeaderItem(colno).text()
                            matches = index.data(Qt.ItemDataRole.UserRole)
                            csv_writer.writerows((filename, structure, match) for match in matches)
            QMessageBox.information(self, "Success", f"Matches has been successfully exported to {file_path}.")
        except PermissionError:
            QMessageBox.critical(
                self, "Error Exporting Cells", f"PermissionError: failed to export matches to {file_path}."
            )
