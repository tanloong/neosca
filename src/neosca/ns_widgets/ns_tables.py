#!/usr/bin/env python3

import os.path as os_path
from typing import Any, Generator, Iterable, List, Literal, Optional, Sequence, Union

from PySide6.QtCore import QModelIndex, QPersistentModelIndex, QSortFilterProxyModel, Signal
from PySide6.QtGui import (
    QStandardItem,
    QStandardItemModel,
    Qt,
)
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QFileDialog,
    QHeaderView,
    QMessageBox,
    QTableView,
    QWidget,
)

from neosca import DESKTOP_PATH
from neosca.ns_qss import Ns_QSS
from neosca.ns_settings.ns_settings import Ns_Settings
from neosca.ns_settings.ns_settings_default import available_export_types
from neosca.ns_widgets.ns_widgets import Ns_MessageBox_Question


class Ns_StandardItemModel(QStandardItemModel):
    data_cleared = Signal()
    row_added = Signal()
    data_exported = Signal()

    def __init__(
        self,
        main,
        hor_labels: Optional[Sequence[str]] = None,
        ver_labels: Optional[Sequence[str]] = None,
        orientation: Literal["hor", "ver"] = "hor",
        show_empty_row: bool = True,
        **kwargs,
    ) -> None:
        super().__init__(main, **kwargs)
        self.main = main
        if hor_labels is not None:
            self.setColumnCount(len(hor_labels))
            self.setHorizontalHeaderLabels(hor_labels)
        if ver_labels is not None:
            self.setRowCount(len(ver_labels))
            self.setVerticalHeaderLabels(ver_labels)

        self.orientation = orientation
        if show_empty_row:
            if orientation == "hor":
                self.setRowCount(1)
            elif orientation == "ver":
                self.setColumnCount(1)
            else:
                assert False, f"Invalid orientation: {orientation}"
        self.show_empty_row = show_empty_row

        self.has_been_exported: bool = False
        self.row_added.connect(lambda: self.set_has_been_exported(False))
        self.data_exported.connect(lambda: self.set_has_been_exported(True))

    def set_item_left_shifted(self, rowno: int, colno: int, value: Union[QStandardItem, str]) -> QStandardItem:
        if isinstance(value, QStandardItem):
            item = value
        elif isinstance(value, str):
            item = QStandardItem()
            # https://stackoverflow.com/a/20469423/20732031
            item.setData(value, Qt.ItemDataRole.DisplayRole)
        else:
            assert False, f"Invalid value type: {type(value)}"
        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.setItem(rowno, colno, item)
        return item

    def set_row_left_shifted(
        self, rowno: int, values: Iterable[Union[QStandardItem, str]], start: int = 0
    ) -> None:
        for colno, value in enumerate(values, start=start):
            self.set_item_left_shifted(rowno, colno, value)

    def set_item_right_shifted(
        self, rowno: int, colno: int, value: Union[QStandardItem, str, int, float]
    ) -> QStandardItem:
        if isinstance(value, QStandardItem):
            item = value
        elif isinstance(value, (str, int, float)):
            item = QStandardItem()
            # https://stackoverflow.com/a/20469423/20732031
            item.setData(value, Qt.ItemDataRole.DisplayRole)
        else:
            assert False, f"Invalid value type: {type(value)}"
        item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.setItem(rowno, colno, item)
        return item

    def set_row_right_shifted(
        self, rowno: int, values: Iterable[Union[QStandardItem, str, int, float]], start: int = 0
    ) -> None:
        for colno, value in enumerate(values, start=start):
            self.set_item_right_shifted(rowno, colno, value)

    def set_has_been_exported(self, exported: bool) -> None:
        self.has_been_exported = exported

    def _clear_data(self) -> None:
        if self.orientation == "hor":
            self.removeRows(0, self.rowCount())
            if self.show_empty_row:
                self.setRowCount(1)
        elif self.orientation == "ver":
            self.setColumnCount(0)
            if self.show_empty_row:
                self.setColumnCount(1)
        self.data_cleared.emit()

    def clear_data(self, confirm=False) -> None:
        """
        Clear data, reserve headers
        """
        if not confirm or self.has_been_exported:
            return self._clear_data()
        else:
            messagebox = Ns_MessageBox_Question(
                self.main,
                "Clear Table",
                "The table has not been exported yet and all the data will be lost. Continue?",
            )
            if messagebox.exec():
                self._clear_data()

    def is_row_empty(self, rowno: int) -> bool:
        for colno in range(self.columnCount()):
            item = self.item(rowno, colno)
            if item is not None and item.text() != "":
                return False
        return True

    def is_empty(self) -> bool:
        return all(self.is_row_empty(rowno) for rowno in range(self.rowCount()))

    def has_user_data(self) -> bool:
        for rowno in range(self.rowCount()):
            for colno in range(self.columnCount()):
                if self.item(rowno, colno).data(Qt.ItemDataRole.UserRole):
                    return True
        return False

    def remove_empty_rows(self) -> None:
        for rowno in reversed(range(self.rowCount())):
            if self.is_row_empty(rowno):
                self.removeRow(rowno)

    # Type hint for generator: https://docs.python.org/3.12/library/typing.html#typing.Generator
    def yield_column(
        self, colno: int, role: Qt.ItemDataRole = Qt.ItemDataRole.DisplayRole
    ) -> Generator[Any, None, None]:
        items = (self.item(rowno, colno) for rowno in range(self.rowCount()))
        return (item.data(role) for item in items if item is not None)

    def yield_checked_item_data(self, colno: int, role: Qt.ItemDataRole) -> Generator[Any, None, None]:
        for rowno in range(self.rowCount()):
            item = self.item(rowno, colno)
            if item is None:
                continue
            if item.checkState() == Qt.CheckState.Checked:
                yield item.data(role)

    # https://stackoverflow.com/questions/75038194/qt6-how-to-disable-selection-for-empty-cells-in-qtableview
    # def flags(self, index) -> Qt.ItemFlag:
    #     if index.data() is None:
    #         return Qt.ItemFlag.NoItemFlags
    #     return super().flags(index)


class Ns_SortFilterProxyModel(QSortFilterProxyModel):
    def __init__(self, main, source_model: Ns_StandardItemModel):
        super().__init__(main)
        self.main = main
        self.source_model = source_model
        self.setSourceModel(source_model)

        self.setDynamicSortFilter(False)

    # Override to sepcify the return type
    def sourceModel(self) -> Ns_StandardItemModel:
        return self.source_model

    # Override
    # https://www.qtcentre.org/threads/22120-No-Sort-Vertical-Header?p=107720#post107720
    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.ItemDataRole.DisplayRole):
        if orientation != Qt.Orientation.Vertical or role != Qt.ItemDataRole.DisplayRole:
            return super().headerData(section, orientation, role)
        else:
            return section + 1


class Ns_TableView(QTableView):
    def __init__(
        self,
        main,
        model: Union[Ns_StandardItemModel, Ns_SortFilterProxyModel],
        has_hor_header: bool = True,
        has_ver_header: bool = False,
        **kwargs,
    ) -> None:
        super().__init__(main, **kwargs)
        self.main = main
        self.setModel(model)
        if isinstance(model, Ns_StandardItemModel):
            self.source_model = model
        elif isinstance(model, Ns_SortFilterProxyModel):
            self.source_model = model.sourceModel()
            self.setSortingEnabled(True)
        else:
            assert False, f"Invalid model type: {model.__class__.__name__}"
        self.source_model.data_cleared.connect(self.on_data_cleared)
        self.source_model.row_added.connect(self.on_row_added)
        self.has_hor_header = has_hor_header
        self.has_ver_header = has_ver_header

        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.horizontalHeader().setHighlightSections(False)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.verticalHeader().setHighlightSections(False)
        self.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

        if self.source_model.is_empty():
            self.setEnabled(False)

    def on_data_cleared(self) -> None:
        self.horizontalHeader().resizeSections()
        self.setEnabled(False)

    def on_row_added(self) -> None:
        if not self.isEnabled():
            self.setEnabled(True)
        self.resizeRowsToContents()
        self.resizeColumnsToContents()

    # Override
    def setIndexWidget(self, index: QModelIndex | QPersistentModelIndex, widget: QWidget) -> None:
        super().setIndexWidget(index, widget)
        if not self.isEnabled():
            self.setEnabled(True)

    def set_openpyxl_horizontal_header_alignment(self, cell) -> None:
        from openpyxl.styles import Alignment

        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def set_openpyxl_vertical_header_alignment(self, cell) -> None:
        from openpyxl.styles import Alignment

        if self.has_ver_header:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def set_openpyxl_cell_alignment(self, cell, item: QStandardItem) -> None:
        # https://doc.qt.io/qtforpython-6/PySide6/QtCore/Qt.html#PySide6.QtCore.PySide6.QtCore.Qt.AlignmentFlag
        # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.alignment.html
        # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L887
        # Qt
        #  Horizontal: Qt.AlignLeft, Qt.AlignRight,  Qt.AlignHCenter, Qt.AlignJustify
        #  Vertical:   Qt.AlignTop,  Qt.AlignBottom, Qt.AlignVCenter, Qt.AlignBaseline
        # OpenPyXL
        #  Horizontal: justify, center, distributed, left, right, fill, general, centerContinuous
        #  Vertical:   justify, center, distributed, top,  bottom

        from openpyxl.styles import Alignment

        alignment_item: Qt.AlignmentFlag = item.textAlignment()

        # Horizontal
        if alignment_item & Qt.AlignmentFlag.AlignLeft:
            alignment_cell_horizontal = "left"
        elif alignment_item & Qt.AlignmentFlag.AlignRight:
            alignment_cell_horizontal = "right"
        elif alignment_item & Qt.AlignmentFlag.AlignHCenter:
            alignment_cell_horizontal = "center"
        elif alignment_item & Qt.AlignmentFlag.AlignJustify:
            alignment_cell_horizontal = "justify"
        else:
            alignment_cell_horizontal = "left"

        # Vertical
        if Qt.AlignmentFlag.AlignTop in alignment_item:
            alignment_cell_vertical = "top"
        elif Qt.AlignmentFlag.AlignBottom in alignment_item:
            alignment_cell_vertical = "bottom"
        elif Qt.AlignmentFlag.AlignVCenter in alignment_item:
            alignment_cell_vertical = "center"
        # > Wordless: Not sure
        elif Qt.AlignmentFlag.AlignBaseline in alignment_item:
            alignment_cell_vertical = "justify"
        else:
            alignment_cell_vertical = "center"

        cell.alignment = Alignment(
            horizontal=alignment_cell_horizontal, vertical=alignment_cell_vertical, wrap_text=True
        )

    def export_table(self, filename: str = "") -> None:
        file_path, file_type = QFileDialog.getSaveFileName(
            parent=self,
            caption="Export Table",
            dir=str(DESKTOP_PATH / filename),
            filter=";;".join(available_export_types),
            selectedFilter=Ns_Settings.value("Export/default-type"),
        )
        if not file_path:
            return

        model = self.model()
        col_count = model.columnCount()
        row_count = model.rowCount()
        try:
            if ".xlsx" in file_type:
                # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L701C1-L716C54
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                from openpyxl.utils import get_column_letter

                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet_cell = worksheet.cell

                rowno_cell_offset = 2 if self.has_hor_header else 1
                colno_cell_offset = 2 if self.has_ver_header else 1

                # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L628C3-L629C82
                dpi_horizontal = QApplication.primaryScreen().logicalDotsPerInchX()
                dpi_vertical = QApplication.primaryScreen().logicalDotsPerInchY()

                font_size = Ns_Settings.value("Appearance/font-size", type=int)

                # 1. Horizontal header text and alignment
                if self.has_hor_header:
                    for colno in range(col_count):
                        cell = worksheet_cell(1, colno_cell_offset + colno)
                        cell.value = model.headerData(colno, Qt.Orientation.Horizontal)
                        self.set_openpyxl_horizontal_header_alignment(cell)
                # 2. Vertical header text and alignment
                if self.has_ver_header:
                    for rowno in range(row_count):
                        cell = worksheet_cell(rowno_cell_offset + rowno, 1)
                        cell.value = model.headerData(rowno, Qt.Orientation.Vertical)
                        self.set_openpyxl_vertical_header_alignment(cell)

                # 3. Both header background and font
                # 3.0.1 Get header background
                horizon_bacolor: Optional[str] = Ns_QSS.get_value(
                    self.main.styleSheet(), "QHeaderView::section:horizontal", "background-color"
                )
                vertical_bacolor: Optional[str] = Ns_QSS.get_value(
                    self.main.styleSheet(), "QHeaderView::section:vertical", "background-color"
                )
                # 3.0.2 Get header font, currently only consider color and boldness
                #  https://www.codespeedy.com/change-font-color-of-excel-cells-using-openpyxl-in-python/
                #  https://doc.qt.io/qt-6/stylesheet-reference.html#font-weight
                header_font_color = Ns_QSS.get_value(self.main.styleSheet(), "QHeaderView::section", "color")
                header_font_color = header_font_color.lstrip("#") if header_font_color is not None else "000"
                header_font_weight = Ns_QSS.get_value(
                    self.main.styleSheet(), "QHeaderView::section", "font-weight"
                )
                header_is_bold = (header_font_weight == "bold") if header_font_weight is not None else False
                # 3.1 Horizontal header background and font
                if self.has_hor_header:
                    # 3.1.1 Horizontal header background
                    #  TODO: Currently all tabpages share the same style sheet and the
                    #   single QSS file is loaded from MainWindow, thus here the
                    #   style sheet is accessed from self. In the future different
                    #   tabs might load their own QSS files, and the style sheet
                    #   should be accessed from the QTabWidget. This is also the
                    #   case for all other "self.styleSheet()" expressions, when
                    #   making this change, remember to edit all of them.
                    if horizon_bacolor is not None:
                        horizon_bacolor = horizon_bacolor.lstrip("#")
                        for colno in range(col_count):
                            cell = worksheet_cell(1, colno_cell_offset + colno)
                            cell.fill = PatternFill(fill_type="solid", fgColor=horizon_bacolor)
                    # 3.1.2 Horizontal header font
                    for colno in range(col_count):
                        cell = worksheet_cell(1, colno_cell_offset + colno)
                        cell.font = Font(
                            color=header_font_color,
                            bold=header_is_bold,
                            size=font_size,
                        )
                # 3.2 Vertical header background and font
                if self.has_ver_header:
                    # 3.2.1 Vertial header background
                    if vertical_bacolor is not None:
                        vertical_bacolor = vertical_bacolor.lstrip("#")
                        for rowno in range(row_count):
                            cell = worksheet_cell(rowno_cell_offset + rowno, 1)
                            cell.fill = PatternFill(fill_type="solid", fgColor=vertical_bacolor)
                    # 3.2.2 Vertical header font
                    for rowno in range(row_count):
                        cell = worksheet_cell(rowno_cell_offset + rowno, 1)
                        cell.font = Font(
                            color=header_font_color,
                            bold=header_is_bold,
                            size=font_size,
                        )

                # 4. Cells
                for rowno in range(row_count):
                    for colno in range(col_count):
                        if isinstance(model, (Ns_SortFilterProxyModel, QSortFilterProxyModel)):
                            mapped_index = model.mapToSource(model.index(rowno, colno))
                            item = mapped_index.model().item(mapped_index.row(), mapped_index.column())
                            item_data = mapped_index.data()
                        else:
                            item = model.item(rowno, colno)
                            item_data = item.data()

                        try:  # noqa: SIM105
                            item_data = float(item_data)
                        except ValueError:
                            pass

                        cell = worksheet_cell(rowno_cell_offset + rowno, colno_cell_offset + colno)
                        cell.value = item_data
                        self.set_openpyxl_cell_alignment(cell, item)
                        cell.font = Font(size=font_size)
                # 5. Column width
                for colno in range(col_count):
                    # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L729
                    worksheet.column_dimensions[get_column_letter(colno_cell_offset + colno)].width = (
                        self.horizontalHeader().sectionSize(colno) / dpi_horizontal * 13 + 3
                    )
                if self.has_ver_header:
                    # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L731
                    worksheet.column_dimensions[get_column_letter(1)].width = (
                        self.verticalHeader().width() / dpi_horizontal * 13 + 3
                    )
                # 6. Row height
                for i, _ in enumerate(worksheet.rows):
                    worksheet.row_dimensions[2 + i].height = (
                        self.verticalHeader().sectionSize(0) / dpi_vertical * 72
                    )
                if self.has_hor_header:
                    worksheet.row_dimensions[1].height = self.horizontalHeader().height() / dpi_vertical * 72

                # 7. Freeze panes
                # https://stackoverflow.com/questions/73837417/freeze-panes-first-two-rows-and-column-with-openpyxl
                # Using "2" in both cases means to always freeze the 1st column
                if self.has_hor_header:
                    worksheet.freeze_panes = "B2"
                else:
                    worksheet.freeze_panes = "A2"
                workbook.save(file_path)
            elif ".csv" in file_type or ".tsv" in file_type:
                import csv

                dialect = csv.excel if ".csv" in file_type else csv.excel_tab
                with open(os_path.normpath(file_path), "w", newline="", encoding="utf-8") as fh:
                    csv_writer = csv.writer(fh, dialect=dialect, lineterminator="\n")
                    # Horizontal header
                    hor_header: List[str] = [""]
                    hor_header.extend(
                        model.headerData(colno, Qt.Orientation.Horizontal) for colno in range(col_count)
                    )
                    csv_writer.writerow(hor_header)
                    # Vertical header + cells
                    if self.has_ver_header:
                        for rowno in range(row_count):
                            row: List[str] = [model.headerData(rowno, Qt.Orientation.Vertical)]
                            row.extend(model.index(rowno, colno).data() for colno in range(col_count))
                            csv_writer.writerow(row)
                    else:
                        for rowno in range(row_count):
                            csv_writer.writerow(model.index(rowno, colno).data() for colno in range(col_count))
            QMessageBox.information(
                self, "Success", f"The table has been successfully exported to {file_path}."
            )
        except PermissionError:
            QMessageBox.critical(
                self, "Error Exporting Cells", f"PermissionError: failed to export the table to {file_path}."
            )
        else:
            self.source_model.data_exported.emit()

    def export_matches(self) -> None:
        file_path, file_type = QFileDialog.getSaveFileName(
            parent=self,
            caption="Export Table",
            dir=str(DESKTOP_PATH / "neosca_sca_matches.xlsx"),
            filter=";;".join(available_export_types),
            selectedFilter=Ns_Settings.value("Export/default-type"),
        )
        if not file_path:
            return

        model = self.model()
        col_count = model.columnCount()
        row_count = model.rowCount()
        try:
            if ".xlsx" in file_type:
                import openpyxl
                from openpyxl.styles import Alignment, Font
                from openpyxl.utils import get_column_letter

                workbook = openpyxl.Workbook()
                for sheetname in workbook.sheetnames:
                    workbook.remove(workbook.get_sheet_by_name(sheetname))

                # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L628C3-L629C82
                dpi_horizontal = QApplication.primaryScreen().logicalDotsPerInchX()
                dpi_vertical = QApplication.primaryScreen().logicalDotsPerInchY()

                font_size = Ns_Settings.value("Appearance/font-size", type=int)

                # Number of matches is much more than that of files. There are
                # 2000 matches across different structures for a 19KB test
                # file. Excel has the max row limit 1,048,576 on a worksheet,
                # which can hold matches of ~524 19KB test files. Here split
                # the matches into different worksheets in case users' corpus
                # is larger than that.
                # https://support.microsoft.com/en-us/office/excel-specifications-and-limits-1672b34d-7043-467e-8e27-269d656771c3
                horizon_header_colwith = 0
                for rowno in range(row_count):
                    if self.has_ver_header:
                        filename = model.headerData(rowno, Qt.Orientation.Vertical)
                    else:
                        filename = model.index(rowno, 0).data()
                    for colno in range(col_count):
                        index = model.index(rowno, colno)
                        if not index.data(Qt.ItemDataRole.UserRole):
                            continue

                        sname = model.headerData(colno, Qt.Orientation.Horizontal)
                        if filename not in workbook.sheetnames:
                            ws = workbook.create_sheet(filename)
                            ws_max_row_offset = 0
                        else:
                            ws = workbook[filename]
                            ws_max_row_offset = 1

                        ws_cell = ws.cell
                        ws_start_rowno = ws.max_row + ws_max_row_offset
                        matches: List[str] = index.data(Qt.ItemDataRole.UserRole)
                        for i, match in enumerate(matches):
                            cell_structure = ws_cell(ws_start_rowno + i, 1)
                            cell_structure.value = sname
                            self.set_openpyxl_horizontal_header_alignment(cell_structure)

                            cell_match = ws_cell(ws_start_rowno + i, 2)
                            cell_match.value = match
                            cell_match.alignment = Alignment(horizontal="left", vertical="center")
                            cell_match.font = Font(size=font_size)

                    horizon_header_colwith = max(
                        horizon_header_colwith, self.horizontalHeader().sectionSize(colno)
                    )

                for ws in workbook.worksheets:
                    # Column width
                    ws.column_dimensions[get_column_letter(1)].width = (
                        horizon_header_colwith / dpi_horizontal * 13 + 3
                    )
                    # Row height
                    for i, _ in enumerate(ws.rows):
                        ws.row_dimensions[1 + i].height = self.horizontalHeader().height() / dpi_vertical * 72

                # Header background color
                # horizon_bacolor: Optional[str] = Ns_QSS.get_value(
                #     self.main.styleSheet(), "QHeaderView::section:horizontal", "background-color"
                # )
                # if horizon_bacolor is not None:
                #     horizon_bacolor = horizon_bacolor.lstrip("#")
                #     for ws in workbook.worksheets:
                #         for cell in ws[get_column_letter(1)]:
                #             cell.fill = PatternFill(fill_type="solid", fgColor=horizon_bacolor)
                vertical_bacolor: Optional[str] = Ns_QSS.get_value(
                    self.main.styleSheet(), "QHeaderView::section:vertical", "background-color"
                )
                if vertical_bacolor is not None:
                    vertical_bacolor = vertical_bacolor.lstrip("#")
                    for ws in workbook.worksheets:
                        # https://openpyxl.readthedocs.io/en/stable/worksheet_properties.html#available-properties-for-worksheets
                        ws.sheet_properties.tabColor = vertical_bacolor
                        ws.sheet_view.showGridLines = False
                # Header font
                for ws in workbook.worksheets:
                    for cell in ws[get_column_letter(1)]:
                        cell.font = Font(size=Ns_Settings.value("Appearance/font-size", type=int))

                # Freeze panes
                for ws in workbook.worksheets:
                    ws.freeze_panes = "B1"

                workbook.save(file_path)
            elif ".csv" in file_type or ".tsv" in file_type:
                import csv

                dialect = csv.excel if ".csv" in file_type else csv.excel_tab
                with open(os_path.normpath(file_path), "w", newline="", encoding="utf-8") as fh:
                    csv_writer = csv.writer(fh, dialect=dialect, lineterminator="\n")
                    header: List[str] = ["Filename", "Structure", "Match"]
                    # instance of "Structure" is stored in "sname"
                    csv_writer.writerow(header)
                    for rowno in range(row_count):
                        if self.has_ver_header:
                            filename = model.headerData(rowno, Qt.Orientation.Vertical)
                        else:
                            filename = model.index(rowno, 0).data()

                        for colno in range(col_count):
                            index = model.index(rowno, colno)
                            if not index.data(Qt.ItemDataRole.UserRole):
                                continue
                            sname = model.headerData(colno, Qt.Orientation.Horizontal)
                            matches = index.data(Qt.ItemDataRole.UserRole)
                            csv_writer.writerows((filename, sname, match) for match in matches)
            QMessageBox.information(self, "Success", f"Matches has been successfully exported to {file_path}.")
        except PermissionError:
            QMessageBox.critical(
                self, "Error Exporting Cells", f"PermissionError: failed to export matches to {file_path}."
            )
