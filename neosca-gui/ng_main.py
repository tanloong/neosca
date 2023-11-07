#!/usr/bin/env python3
# -*- coding=utf-8 -*-

import os
import os.path as os_path
import subprocess
import sys
from typing import Dict, Iterable, List, Literal, Optional, Set

from PySide6.QtCore import QModelIndex, QObject, QThread, Qt, Signal
from PySide6.QtGui import (
    QAction,
    QCursor,
    QKeySequence,
    QPalette,
    QStandardItem,
    QStandardItemModel,
)
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QDialog,
    QFileDialog,
    QGridLayout,
    QGroupBox,
    QHeaderView,
    QLabel,
    QMainWindow,
    QMenu,
    QMenuBar,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QSplitter,
    QTabWidget,
    QTableView,
    QWidget,
)
from .neosca.lca.lca import LCA
from .neosca.neosca import NeoSCA
from .neosca.structure_counter import StructureCounter


class Ng_Main(QMainWindow):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setup_env()

        self.setWindowTitle("NeoSCA-GUI")
        self.setup_menu()
        self.setup_worker()
        self.setup_main_window()

    def setup_menu(self):
        menubar = QMenuBar()
        # menubar.setStyleSheet("background-color: cyan;")
        menu_file = QMenu("File")
        action_open_file = QAction("Open File...", menu_file)
        action_open_file.setShortcut(QKeySequence(Qt.CTRL | Qt.Key_O))
        action_open_file.triggered.connect(self.browse_file)
        action_open_folder = QAction("Open Folder...", menu_file)
        action_restart = QAction("Restart", menu_file)  # TODO remove this before releasing
        action_restart.triggered.connect(self.restart)  # TODO remove this before releasing
        action_restart.setShortcut(
            QKeySequence(Qt.CTRL | Qt.Key_R)
        )  # TODO remove this before releasing
        action_quit = QAction("Quit", menu_file)
        action_quit.setShortcut(QKeySequence(Qt.CTRL | Qt.Key_Q))
        action_quit.triggered.connect(self.close)
        menu_file.addAction(action_open_file)
        menu_file.addAction(action_open_folder)
        menu_file.addAction(action_restart)
        menu_file.addAction(action_quit)
        menubar.addMenu(menu_file)
        self.setMenuBar(menubar)

    def setup_tab_sca(self):
        # frame_preview.setStyleSheet("background-color: green;")
        self.model_sca = QStandardItemModel()
        self.model_sca.setColumnCount(len(StructureCounter.DEFAULT_MEASURES))
        self.model_sca.setHorizontalHeaderLabels(StructureCounter.DEFAULT_MEASURES)
        self.reset_model(self.model_sca, orientation="hor")
        self.tableview_preview_sca = QTableView()
        self.tableview_preview_sca.setModel(self.model_sca)
        # self.table_preview_sca.setStyleSheet("background-color: #C7C7C7;")
        self.tableview_preview_sca.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tableview_preview_sca.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )
        self.tableview_preview_sca.horizontalHeader().setHighlightSections(False)
        self.tableview_preview_sca.verticalHeader().setHighlightSections(False)
        # self.tableview_preview_sca.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.tableview_preview_sca.setSelectionBehavior(QTableView.SelectRows)

        self.button_generate_table_sca = QPushButton("Generate table")
        self.button_generate_table_sca.setEnabled(False)
        self.button_generate_table_sca.clicked.connect(self.ng_thread_sca_generate_table.start)
        self.button_generate_table_sca.setShortcut(QKeySequence(Qt.CTRL | Qt.Key_G))
        self.button_export_table_sca = QPushButton("Export all cells...")
        self.button_export_table_sca.setEnabled(False)
        self.button_export_table_sca.clicked.connect(lambda: self.export_table(self.model_sca))
        # self.button_export_selected_cells = QPushButton("Export selected cells...")
        # self.button_export_selected_cells.setEnabled(False)

        # frame_setting_sca.setStyleSheet("background-color: pink;")
        self.checkbox_reserve_parsed_trees = QCheckBox(
            "Reserve parsed trees",
        )
        self.checkbox_reserve_parsed_trees.setChecked(True)
        self.checkbox_reserve_matched_subtrees = QCheckBox("Reserve matched subtrees")
        self.checkbox_reserve_matched_subtrees.setChecked(True)
        widget_settings_sca = QWidget()
        widget_settings_sca.setLayout(QGridLayout())
        widget_settings_sca.layout().addWidget(self.checkbox_reserve_parsed_trees, 0, 0)
        widget_settings_sca.layout().addWidget(self.checkbox_reserve_matched_subtrees, 1, 0)

        scrollarea_settings_sca = QScrollArea()
        scrollarea_settings_sca.setLayout(QGridLayout())
        scrollarea_settings_sca.setWidgetResizable(True)
        scrollarea_settings_sca.setFixedWidth(200)
        scrollarea_settings_sca.setBackgroundRole(QPalette.Light)
        scrollarea_settings_sca.setWidget(widget_settings_sca)

        self.tab_sca = QWidget()
        self.tab_sca.setLayout(QGridLayout())
        self.tab_sca.layout().addWidget(self.tableview_preview_sca, 0, 0, 1, 2)
        self.tab_sca.layout().addWidget(self.button_generate_table_sca, 1, 0)
        self.tab_sca.layout().addWidget(self.button_export_table_sca, 1, 1)
        self.tab_sca.layout().addWidget(scrollarea_settings_sca, 0, 2, 2, 1)
        self.tab_sca.layout().setContentsMargins(6, 4, 6, 4)

    def setup_tab_lca(self):
        # frame_preview.setStyleSheet("background-color: green;")
        self.model_lca = QStandardItemModel()
        self.model_lca.setColumnCount(len(LCA.FIELDNAMES) - 1)
        self.model_lca.setHorizontalHeaderLabels(LCA.FIELDNAMES[1:])
        self.reset_model(self.model_lca, orientation="hor")
        self.tableview_preview_lca = QTableView()
        self.tableview_preview_lca.setModel(self.model_lca)
        self.tableview_preview_lca.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tableview_preview_lca.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )
        self.tableview_preview_lca.horizontalHeader().setHighlightSections(False)
        self.tableview_preview_lca.verticalHeader().setHighlightSections(False)
        # self.tableview_preview_lca.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.tableview_preview_lca.setSelectionBehavior(QTableView.SelectRows)

        self.button_generate_table_lca = QPushButton("Generate table")
        self.button_generate_table_lca.setEnabled(False)
        self.button_generate_table_lca.clicked.connect(self.ng_thread_lca_generate_table.start)
        self.button_generate_table_lca.setShortcut(QKeySequence(Qt.CTRL | Qt.Key_G))
        self.button_export_table_lca = QPushButton("Export all cells...")
        self.button_export_table_lca.setEnabled(False)
        self.button_export_table_lca.clicked.connect(lambda: self.export_table(self.model_lca))
        # self.button_export_selected_cells = QPushButton("Export selected cells...")
        # self.button_export_selected_cells.setEnabled(False)

        self.radiobutton_wordlist_BNC = QRadioButton("British National Corpus (BNC) wordlist")
        self.radiobutton_wordlist_BNC.setChecked(True)
        self.radiobutton_wordlist_ANC = QRadioButton("American National Corpus (ANC) wordlist")
        groupbox_wordlist = QGroupBox("Wordlist")
        groupbox_wordlist.setLayout(QGridLayout())
        groupbox_wordlist.layout().addWidget(self.radiobutton_wordlist_BNC, 0, 0)
        groupbox_wordlist.layout().addWidget(self.radiobutton_wordlist_ANC, 1, 0)
        self.radiobutton_tagset_ud = QRadioButton("Universal POS Tagset")
        self.radiobutton_tagset_ud.setChecked(True)
        self.radiobutton_tagset_ptb = QRadioButton("Penn Treebank POS Tagset")
        groupbox_tagset = QGroupBox("Tagset")
        groupbox_tagset.setLayout(QGridLayout())
        groupbox_tagset.layout().addWidget(self.radiobutton_tagset_ud, 0, 0)
        groupbox_tagset.layout().addWidget(self.radiobutton_tagset_ptb, 1, 0)

        widget_settings_lca = QWidget()
        widget_settings_lca.setLayout(QGridLayout())
        widget_settings_lca.layout().addWidget(groupbox_wordlist, 0, 0)
        widget_settings_lca.layout().addWidget(groupbox_tagset, 1, 0)

        scrollarea_settings_lca = QScrollArea()
        scrollarea_settings_lca.setFixedWidth(200)
        scrollarea_settings_lca.setWidgetResizable(True)
        scrollarea_settings_lca.setBackgroundRole(QPalette.Light)
        scrollarea_settings_lca.setWidget(widget_settings_lca)

        self.tab_lca = QWidget()
        self.tab_lca.setLayout(QGridLayout())
        self.tab_lca.layout().addWidget(self.tableview_preview_lca, 0, 0, 1, 2)
        self.tab_lca.layout().addWidget(self.button_generate_table_lca, 1, 0)
        self.tab_lca.layout().addWidget(self.button_export_table_lca, 1, 1)
        self.tab_lca.layout().addWidget(scrollarea_settings_lca, 0, 2, 2, 1)
        self.tab_lca.layout().setContentsMargins(6, 4, 6, 4)

    def setup_main_window(self):
        self.model_file = QStandardItemModel()
        self.model_file.setHorizontalHeaderLabels(("Name", "Path"))
        self.reset_model(self.model_file)
        self.tableview_file = QTableView()
        self.tableview_file.setModel(self.model_file)
        self.tableview_file.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tableview_file.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )
        self.tableview_file.horizontalHeader().setHighlightSections(False)
        self.tableview_file.verticalHeader().setHighlightSections(False)
        # self.tableview_file.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.tableview_file.setSelectionBehavior(QTableView.SelectRows)
        self.tableview_file.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        # https://doc.qt.io/qtforpython-6/PySide6/QtWidgets/QWidget.html#PySide6.QtWidgets.PySide6.QtWidgets.QWidget.customContextMenuRequested
        self.tableview_file.customContextMenuRequested.connect(self.show_menu_for_tableview_file)

        self.setup_tab_sca()
        self.setup_tab_lca()
        self.tabwidget = QTabWidget()
        self.tabwidget.addTab(self.tab_sca, "Syntactic Complexity Analyzer")
        self.tabwidget.addTab(self.tab_lca, "Lexical Complexity Analyzer")
        self.splitter_central_widget = QSplitter(Qt.Orientation.Vertical)
        self.splitter_central_widget.setChildrenCollapsible(False)
        self.splitter_central_widget.addWidget(self.tabwidget)
        self.splitter_central_widget.setStretchFactor(0, 2)
        self.splitter_central_widget.addWidget(self.tableview_file)
        self.splitter_central_widget.setStretchFactor(1, 1)
        self.setCentralWidget(self.splitter_central_widget)

    def setup_worker(self) -> None:
        self.processing_dialog = QDialog(self)
        self.processing_dialog.setWindowTitle("Please waite.")
        self.processing_dialog.resize(300, 200)
        QLabel(
            "NeoSCA is running. It may take a few minutes to finish the job. Please wait.",
            self.processing_dialog,
        )

        self.ng_worker_sca_generate_table = Ng_Worker_SCA_Generate_Table(
            self, self.processing_dialog
        )
        self.ng_thread_sca_generate_table = Ng_Thread(self.ng_worker_sca_generate_table)

        self.ng_worker_lca_generate_table = Ng_Worker_LCA_Generate_Table(
            self, self.processing_dialog
        )
        self.ng_thread_lca_generate_table = Ng_Thread(self.ng_worker_lca_generate_table)

    def reset_model(
        self, model: QStandardItemModel, orientation: Literal["hor", "ver"] = "hor"
    ) -> None:
        # https://stackoverflow.com/questions/75038194/qt6-how-to-disable-selection-for-empty-cells-in-qtableview
        if orientation == "hor":
            model.setRowCount(0)
            model.setRowCount(1)
        elif orientation == "ver":
            model.setColumnCount(0)
            model.setColumnCount(1)

    def resize_tableview(
        self, tableview: QTableView, orientation: Literal["hor", "ver"] = "hor"
    ) -> None:
        model = tableview.model()
        if (
            model.rowCount() >= 1
        ):  # TODO: here need to change when transpose the table to vertical
            tableview.horizontalHeader().setSectionResizeMode(
                QHeaderView.ResizeMode.ResizeToContents
            )
        else:
            self.reset_model(model, orientation)

    def setup_env(self) -> None:
        self.desktop = os_path.normpath(os_path.expanduser("~/Desktop"))
        self.here = os_path.dirname(os_path.abspath(__file__))
        ng_home = os_path.dirname(self.here)
        libs_dir = os_path.join(ng_home, "libs")
        self.java_home = os_path.join(libs_dir, "jdk8u372")
        self.stanford_parser_home = os_path.join(libs_dir, "stanford-parser-full-2020-11-17")
        self.stanford_tregex_home = os_path.join(libs_dir, "stanford-tregex-2020-11-17")
        os.environ["JAVA_HOME"] = self.java_home
        os.environ["STANFORD_PARSER_HOME"] = self.stanford_parser_home
        os.environ["STANFORD_TREGEX_HOME"] = self.stanford_tregex_home
        self.env = os.environ.copy()

    def export_table(self, model: QStandardItemModel) -> None:
        file_path, file_type = QFileDialog.getSaveFileName(
            parent=self,
            caption="Export Table",
            dir=self.desktop,
            filter="Excel Workbook (*.xlsx);;CSV File (*.csv);;TSV File (*.tsv)",
        )
        if not file_path:
            return

        col_count = model.columnCount()
        row_count = model.rowCount()
        try:
            if ".xlsx" in file_type:
                # https://github.com/BLKSerene/Wordless/blob/main/wordless/wl_widgets/wl_tables.py#L701C1-L716C54
                import openpyxl

                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                # Horizontal header
                for colno_cell, colno_item in enumerate(range(col_count)):
                    cell = worksheet.cell(1, 2 + colno_cell)
                    cell.value = model.horizontalHeaderItem(colno_item).text()
                # Vertical header
                for rowno_cell, rowno_item in enumerate(range(row_count)):
                    cell = worksheet.cell(2 + rowno_cell, 1)
                    cell.value = model.verticalHeaderItem(rowno_item).text()
                # Cells
                for rowno_cell, rowno in enumerate(range(row_count)):
                    for colno_cell, colno_item in enumerate(range(col_count)):
                        cell = worksheet.cell(2 + rowno_cell, 2 + colno_cell)
                        cell.value = model.item(rowno, colno_item).text()
                workbook.save(file_path)
            elif ".csv" in file_type or ".tsv" in file_type:
                import csv

                dialect = csv.excel if ".csv" in file_type else csv.excel_tab
                with open(os_path.normpath(file_path), "w", newline="", encoding="utf-8") as fh:
                    csv_writer = csv.writer(fh, dialect=dialect, lineterminator="\n")
                    # Horizontal header
                    hor_header: List[str] = [""]
                    hor_header.extend(
                        model.horizontalHeaderItem(colno).text() for colno in range(col_count)
                    )
                    csv_writer.writerow(hor_header)
                    # Vertical header + cells
                    for rowno in range(row_count):
                        row: List[str] = [model.verticalHeaderItem(rowno).text()]
                        row.extend(model.item(rowno, colno).text() for colno in range(col_count))
                        csv_writer.writerow(row)
            QMessageBox.information(
                self, "Success", f"The table has been successfully exported to {file_path}."
            )
        except PermissionError:
            QMessageBox.critical(
                self,
                "Error Exporting Cells",
                f"PermissionError: failed to export the table to {file_path}.",
            )

    def show_menu_for_tableview_file(self) -> None:
        action_remove_file = QAction("Remove")
        action_remove_file.triggered.connect(self.remove_file_paths)
        menu = QMenu()
        menu.addAction(action_remove_file)
        menu.exec(QCursor.pos())

    def setenable_button_generate_table(self, enabled: bool = False) -> None:
        self.button_generate_table_sca.setEnabled(enabled)
        self.button_generate_table_lca.setEnabled(enabled)

    def remove_file_paths(self) -> None:
        # https://stackoverflow.com/questions/5927499/how-to-get-selected-rows-in-qtableview
        indexes: List[QModelIndex] = self.tableview_file.selectionModel().selectedRows()
        # Remove rows from bottom to top, or otherwise the lower row indexes will change as upper rows are removed
        rownos = sorted((index.row() for index in indexes), reverse=True)
        for rowno in rownos:
            self.model_file.takeRow(rowno)
        if self.model_file.rowCount() == 0:
            self.setenable_button_generate_table(False)

    def remove_model_rows(self, model: QStandardItemModel, *rownos: int) -> None:
        if not rownos:
            # https://doc.qt.io/qtforpython-6/PySide6/QtGui/QStandardItemModel.html#PySide6.QtGui.PySide6.QtGui.QStandardItemModel.setRowCount
            model.setRowCount(0)
        else:
            for rowno in rownos:
                model.takeRow(rowno)

    def remove_model_single_empty_row(self, model: QStandardItemModel) -> None:
        if model.rowCount() == 1 and model.item(0, 0) is None:
            model.setRowCount(0)

    # Type hint for generator: https://docs.python.org/3.12/library/typing.html#typing.Generator
    def yield_model_column(self, model: QStandardItemModel, colno: int) -> Iterable[str]:
        items = (model.item(rowno, colno) for rowno in range(model.rowCount()))
        return (item.text() for item in items if item is not None)

    def yield_added_file_names(self) -> Iterable[str]:
        colno_path = 0
        return self.yield_model_column(self.model_file, colno_path)

    def yield_added_file_paths(self) -> Iterable[str]:
        colno_path = 1
        return self.yield_model_column(self.model_file, colno_path)

    def set_model_items_from_row_start(
        self, model: QStandardItemModel, rowno: int, *items: str
    ) -> None:
        for colno, item in enumerate(items):
            model.setItem(rowno, colno, QStandardItem(item))

    def add_file_paths(self, file_paths_to_add: List[str]) -> None:
        unique_file_paths_to_add: Set[str] = set(file_paths_to_add)
        already_added_file_paths: Set[str] = set(self.yield_added_file_paths())
        file_paths_dup: Set[str] = unique_file_paths_to_add & already_added_file_paths
        file_paths_empty: Set[str] = set(
            filter(lambda p: not os_path.getsize(p), unique_file_paths_to_add)
        )
        file_paths_ok: Set[str] = (
            unique_file_paths_to_add
            - already_added_file_paths
            - file_paths_dup
            - file_paths_empty
        )
        if file_paths_ok:
            self.remove_model_single_empty_row(self.model_file)
            colno_name = 0
            already_added_file_names = list(
                self.yield_model_column(self.model_file, colno_name)
            )  # Here the already_added_file_names will have no duplicates
            for file_path in file_paths_ok:
                file_name = os_path.splitext(os_path.basename(file_path))[0]
                if file_name in already_added_file_names:
                    occurrence = 2
                    while f"{file_name} ({occurrence})" in already_added_file_names:
                        occurrence += 1
                    file_name = f"{file_name} ({occurrence})"
                already_added_file_names.append(file_name)
                self.set_model_items_from_row_start(
                    self.model_file, self.model_file.rowCount(), file_name, file_path
                )

            self.tableview_file.horizontalHeader().setSectionResizeMode(
                QHeaderView.ResizeMode.ResizeToContents
            )
            # Enable "generate_table" button when new files are added
            self.setenable_button_generate_table(True)

        if any((file_paths_dup, file_paths_empty)):  # TODO should show a table
            QMessageBox.information(
                self,
                "Error Adding Files",
                "These files are skipped:\n- {}".format(
                    "\n- ".join(file_paths_dup | file_paths_empty)
                ),
            )

    def browse_file(self):
        file_dialog = QFileDialog(
            directory="/home/tan/docx/corpus/YuHua-parallel-corpus-zh-en/02aligned/standalone/"
        )  # TODO remove this before releasing
        file_paths_to_add, _ = file_dialog.getOpenFileNames(
            self, "Open Files", "", "All files (*.*);;Text files (*.txt);;Docx files (*.docx)"
        )
        if not file_paths_to_add:
            return
        self.add_file_paths(file_paths_to_add)

    def restart(self):
        self.close()
        command = [sys.executable] + sys.argv
        subprocess.call(command, env=os.environ.copy(), close_fds=False)


class Ng_Worker(QObject):
    worker_done = Signal()

    def __init__(self, main, dialog) -> None:
        super().__init__()
        self.main = main
        self.dialog = dialog

    # TODO
    # def cancel(self) -> None:
    #     self.terminate()
    #     self.wait() # TODO ?
    #     <<>>

    def run(self) -> None:
        raise NotImplementedError()


class Ng_Worker_SCA_Generate_Table(Ng_Worker):
    def __init__(self, main, dialog) -> None:
        super().__init__(main, dialog)
        self.worker_done.connect(
            lambda: self.main.resize_tableview(self.main.tableview_preview_sca)
        )
        self.worker_done.connect(lambda: self.main.button_export_table_sca.setEnabled(True))

    def run(self) -> None:
        self.main.button_generate_table_sca.setEnabled(False)
        input_file_names = self.main.yield_added_file_names()
        input_file_paths = self.main.yield_added_file_paths()
        if not input_file_paths:
            QMessageBox.warning(self.main, "No Input Files", f"Please select files to begin.")
            return

        sca_kwargs = {
            "is_auto_save": False,
            "odir_matched": "",
            "selected_measures": None,
            "is_reserve_parsed": self.main.checkbox_reserve_parsed_trees.isChecked(),
            "is_reserve_matched": self.main.checkbox_reserve_matched_subtrees.isChecked(),
            "is_skip_querying": False,
            "is_skip_parsing": False,
            "config": None,
        }

        attrname = "sca_analyzer"
        try:
            sca_analyzer = getattr(self.main, attrname)
        except AttributeError:
            sca_analyzer = NeoSCA(**sca_kwargs)
            setattr(self.main, attrname, sca_analyzer)
        else:
            sca_analyzer.update_options(sca_kwargs)

        self.main.remove_model_rows(self.main.model_sca)
        err_file_paths = []
        for rowno, (file_name, file_path) in enumerate(zip(input_file_names, input_file_paths)):
            try:
                counter: Optional[StructureCounter] = sca_analyzer.parse_and_query_ifile(
                    file_path
                )  # TODO should concern --no-parse, --no-query, ... after adding all available options
            except:
                err_file_paths.append(file_path)
                continue
            if counter is None:
                err_file_paths.append(file_path)
                continue
            sname_value_map: Dict[str, str] = counter.get_all_values()
            _, *items = (QStandardItem(value) for value in sname_value_map.values())
            self.main.model_sca.insertRow(rowno, items)
            self.main.model_sca.setVerticalHeaderItem(rowno, QStandardItem(file_name))

        if err_file_paths:  # TODO: should show a table
            QMessageBox.information(
                self.main,
                "Error Processing Files",
                "These files are skipped:\n- {}".format("\n- ".join(err_file_paths)),
            )
        self.worker_done.emit()


class Ng_Worker_LCA_Generate_Table(Ng_Worker):
    def __init__(self, main, dialog) -> None:
        super().__init__(main, dialog)
        self.worker_done.connect(
            lambda: self.main.resize_tableview(self.main.tableview_preview_lca)
        )
        self.worker_done.connect(lambda: self.main.button_export_table_lca.setEnabled(True))

    def run(self) -> None:
        self.main.button_generate_table_lca.setEnabled(False)
        input_file_names = self.main.yield_added_file_names()
        input_file_paths = self.main.yield_added_file_paths()
        if not input_file_paths:
            QMessageBox.warning(self.main, "No input files", f"Please select files to process.")
            return

        lca_kwargs = {
            "wordlist": "bnc" if self.main.radiobutton_wordlist_BNC.isChecked() else "anc",
            "tagset": "ud" if self.main.radiobutton_tagset_ud.isChecked() else "ptb",
            "is_stdout": False,
        }
        attrname = "lca_analyzer"
        try:
            lca_analyzer = getattr(self.main, attrname)
        except AttributeError:
            lca_analyzer = LCA(**lca_kwargs)
            setattr(self.main, attrname, lca_analyzer)
        else:
            lca_analyzer.update_options(lca_kwargs)

        self.main.remove_model_rows(self.main.model_lca)
        err_file_paths = []
        for rowno, (file_name, file_path) in enumerate(zip(input_file_names, input_file_paths)):
            try:
                values = lca_analyzer._analyze(file_path=file_path)
            except:
                err_file_paths.append(file_path)
                continue
            if values is None:  # TODO: should pop up warning window
                err_file_paths.append(file_path)
                continue
            _, *items = (QStandardItem(value) for value in values)
            self.main.model_lca.insertRow(rowno, items)
            self.main.model_lca.setVerticalHeaderItem(rowno, QStandardItem(file_name))

        if err_file_paths:  # TODO: should show a table
            QMessageBox.information(
                self.main,
                "Error Processing Files",
                "These files are skipped:\n- {}".format("\n- ".join(err_file_paths)),
            )
        self.worker_done.emit()


class Ng_Thread(QThread):
    def __init__(self, worker):
        super().__init__()
        self.worker = worker
        # https://mayaposch.wordpress.com/2011/11/01/how-to-really-truly-use-qthreads-the-full-explanation/
        self.worker.moveToThread(self)

        self.started.connect(self.worker.dialog.exec)
        self.finished.connect(self.worker.dialog.accept)

    def run(self):
        self.start()
        self.worker.run()


class Ng_QSS_Loader:
    def __init__(self):
        pass

    @staticmethod
    def read_qss_file(qss_file_path):
        with open(qss_file_path, "r", encoding="utf-8") as file:
            return file.read()


if __name__ == "__main__":
    ng_app = QApplication(sys.argv)
    ng_window = Ng_Main()
    ng_window.setStyleSheet(
        Ng_QSS_Loader.read_qss_file(os_path.join(ng_window.here, "ng_style.qss"))
    )
    ng_window.show()
    sys.exit(ng_app.exec())
